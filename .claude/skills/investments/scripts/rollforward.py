#!/usr/bin/env python3
"""
Securities Unrealized Gain/Loss Rollforward Report
Generates a rollforward showing movement in unrealized gains/losses by security.
"""

import argparse
import json
import os
import sys
from datetime import datetime
from collections import defaultdict

# Add shared module to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LIB_PATH = os.path.join(SCRIPT_DIR, '..', '..', '..', '..', 'lib')
sys.path.insert(0, LIB_PATH)

from utils import (sanitize_filename, OPENPYXL_AVAILABLE, calculate_period_begin,
                   get_period_dates, get_period_label, VALID_PERIOD_TYPES,
                   get_or_create_run_timestamp, save_journal_csv, save_journal_xlsx)
from accounts import SECURITIES_ACCOUNTS, get_securities_account

if OPENPYXL_AVAILABLE:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

# Import xlsx styles for color-coding
try:
    from xlsx_styles import input_style, formula_style
    STYLES_AVAILABLE = True
except ImportError:
    STYLES_AVAILABLE = False


def parse_args():
    parser = argparse.ArgumentParser(description='Generate securities unrealized G/L rollforward')
    parser.add_argument('--report-date', required=True,
                        help='Report date (YYYY-MM-DD) - period end date')
    parser.add_argument('--period-begin', required=False,
                        help='Period begin date (YYYY-MM-DD). Required unless --period-type is provided.')
    parser.add_argument('--period-type', required=False, choices=VALID_PERIOD_TYPES,
                        help='Period type - calculates begin date from report-date. Required unless --period-begin is provided.')
    parser.add_argument('--output-dir', default='data/output/investments',
                        help='Output directory')
    parser.add_argument('--run-timestamp', default=None,
                        help='Timestamp for output folder (YYYY-MM-DD_HH-MM). If not provided, uses current time.')

    args = parser.parse_args()

    # Validate: must have either period-begin or period-type
    if not args.period_begin and not args.period_type:
        parser.error("Either --period-begin or --period-type is required.")

    return args


def load_securities_register(output_dir):
    """Load securities register."""
    register_path = os.path.join(output_dir, 'securities_register.json')
    if os.path.exists(register_path):
        with open(register_path, 'r') as f:
            return json.load(f)
    return {'transactions': [], 'holdings': {}, 'fmv_adjustments': [], 'metadata': {}}


def calculate_rollforward_data(register, report_date, period_begin):
    """Calculate rollforward data for each security."""
    report_dt = datetime.strptime(report_date, '%Y-%m-%d')
    period_begin_dt = datetime.strptime(period_begin, '%Y-%m-%d')

    # Track data by security
    securities = defaultdict(lambda: {
        'security_type': 'AFS',
        'beginning_cost': 0,
        'beginning_fmv_adj': 0,
        'purchases': 0,
        'sales_cost_basis': 0,
        'sales_proceeds': 0,
        'sales_fmv_adj_release': 0,
        'realized_gain_loss': 0,
        'period_fmv_adj': 0,
        'ending_cost': 0,
        'ending_fmv_adj': 0,
    })

    # First pass: Calculate beginning balances and period activity
    for txn in register.get('transactions', []):
        txn_dt = datetime.strptime(txn['date'], '%Y-%m-%d')
        security = txn['security']
        securities[security]['security_type'] = txn.get('security_type', 'AFS')

        if txn['type'] == 'purchase':
            if txn_dt < period_begin_dt:
                securities[security]['beginning_cost'] += txn['total_cost']
            elif txn_dt <= report_dt:
                securities[security]['purchases'] += txn['total_cost']

        elif txn['type'] == 'sale':
            if txn_dt < period_begin_dt:
                securities[security]['beginning_cost'] -= txn['cost_basis']
            elif txn_dt <= report_dt:
                securities[security]['sales_cost_basis'] += txn['cost_basis']
                proceeds = txn.get('proceeds', txn['cost_basis'])
                securities[security]['sales_proceeds'] += proceeds
                securities[security]['realized_gain_loss'] += (proceeds - txn['cost_basis'])

    # Process FMV adjustments
    for adj in register.get('fmv_adjustments', []):
        adj_dt = datetime.strptime(adj['date'], '%Y-%m-%d')
        security = adj['security']
        securities[security]['security_type'] = adj.get('security_type', 'AFS')

        if adj_dt < period_begin_dt:
            securities[security]['beginning_fmv_adj'] += adj['adjustment']
        elif adj_dt <= report_dt:
            securities[security]['period_fmv_adj'] += adj['adjustment']

    # Calculate ending balances
    for security, data in securities.items():
        data['ending_cost'] = (
            data['beginning_cost'] +
            data['purchases'] -
            data['sales_cost_basis']
        )

        # Handle unrealized G/L release on sales (reclassification)
        if data['sales_cost_basis'] > 0 and data['beginning_cost'] > 0:
            total_cost_before_sale = data['beginning_cost'] + data['purchases']
            if total_cost_before_sale > 0:
                sale_proportion = data['sales_cost_basis'] / total_cost_before_sale
                unrealized_before_sale = data['beginning_fmv_adj'] + data['period_fmv_adj']
                data['sales_fmv_adj_release'] = unrealized_before_sale * sale_proportion

        data['ending_fmv_adj'] = (
            data['beginning_fmv_adj'] +
            data['period_fmv_adj'] -
            data['sales_fmv_adj_release']
        )

        if data['ending_cost'] <= 0:
            data['ending_cost'] = 0
            data['ending_fmv_adj'] = 0

    return dict(securities)


def generate_journal_entries(register, period_begin, report_date):
    """Generate journal entries from register transactions within period.

    Args:
        register: Securities register with transactions and fmv_adjustments
        period_begin: Period start date (YYYY-MM-DD)
        report_date: Period end date (YYYY-MM-DD)

    Returns:
        List of journal entry dicts suitable for save_journal_csv/xlsx
    """
    entries = []
    period_begin_dt = datetime.strptime(period_begin, '%Y-%m-%d')
    report_dt = datetime.strptime(report_date, '%Y-%m-%d')

    # Process purchases and sales within period
    for txn in register.get('transactions', []):
        txn_dt = datetime.strptime(txn['date'], '%Y-%m-%d')
        if txn_dt < period_begin_dt or txn_dt > report_dt:
            continue

        security = txn['security']
        security_type = txn.get('security_type', 'AFS')
        inv_account = get_securities_account(security_type)
        cash_account = SECURITIES_ACCOUNTS['cash']

        if txn['type'] == 'purchase':
            shares = txn['shares']
            price = txn['price_per_share']
            total = txn['total_cost']
            desc = f"Purchase {shares} shares of {security} @ ${price:,.2f}"

            # DR: Investments
            entries.append({
                'date': txn['date'],
                'type': 'Securities Purchase',
                'asset_name': security,
                'account_code': inv_account['code'],
                'account_name': inv_account['name'],
                'description': desc,
                'debit': total,
                'credit': 0.0
            })
            # CR: Cash
            entries.append({
                'date': txn['date'],
                'type': 'Securities Purchase',
                'asset_name': security,
                'account_code': cash_account['code'],
                'account_name': cash_account['name'],
                'description': desc,
                'debit': 0.0,
                'credit': total
            })

        elif txn['type'] == 'sale':
            shares = txn['shares']
            cost_basis = txn['cost_basis']
            proceeds = txn.get('proceeds', cost_basis)
            gain_loss = txn.get('gain_loss', proceeds - cost_basis)
            desc = f"Sale of {shares} shares of {security}"

            # DR: Cash
            entries.append({
                'date': txn['date'],
                'type': 'Securities Sale',
                'asset_name': security,
                'account_code': cash_account['code'],
                'account_name': cash_account['name'],
                'description': desc,
                'debit': proceeds,
                'credit': 0.0
            })
            # CR: Investments (at cost basis)
            entries.append({
                'date': txn['date'],
                'type': 'Securities Sale',
                'asset_name': security,
                'account_code': inv_account['code'],
                'account_name': inv_account['name'],
                'description': desc,
                'debit': 0.0,
                'credit': cost_basis
            })
            # Gain or Loss
            if gain_loss > 0:
                gain_account = SECURITIES_ACCOUNTS['realized_gain']
                entries.append({
                    'date': txn['date'],
                    'type': 'Securities Sale',
                    'asset_name': security,
                    'account_code': gain_account['code'],
                    'account_name': gain_account['name'],
                    'description': f"Realized gain - {security}",
                    'debit': 0.0,
                    'credit': gain_loss
                })
            elif gain_loss < 0:
                loss_account = SECURITIES_ACCOUNTS['realized_loss']
                entries.append({
                    'date': txn['date'],
                    'type': 'Securities Sale',
                    'asset_name': security,
                    'account_code': loss_account['code'],
                    'account_name': loss_account['name'],
                    'description': f"Realized loss - {security}",
                    'debit': abs(gain_loss),
                    'credit': 0.0
                })

    # Process FMV adjustments at report date
    for adj in register.get('fmv_adjustments', []):
        adj_dt = datetime.strptime(adj['date'], '%Y-%m-%d')
        if adj_dt < period_begin_dt or adj_dt > report_dt:
            continue

        security = adj['security']
        security_type = adj.get('security_type', 'AFS')
        adjustment = adj['adjustment']
        inv_account = get_securities_account(security_type)
        desc = f"FMV adjustment - {security}"

        if security_type == 'AFS':
            # AFS: Unrealized goes to OCI
            oci_account = SECURITIES_ACCOUNTS['oci']
            if adjustment > 0:
                # Write-up: DR Investment, CR OCI
                entries.append({
                    'date': adj['date'],
                    'type': 'FMV Adjustment',
                    'asset_name': security,
                    'account_code': inv_account['code'],
                    'account_name': inv_account['name'],
                    'description': desc,
                    'debit': adjustment,
                    'credit': 0.0
                })
                entries.append({
                    'date': adj['date'],
                    'type': 'FMV Adjustment',
                    'asset_name': security,
                    'account_code': oci_account['code'],
                    'account_name': oci_account['name'],
                    'description': desc,
                    'debit': 0.0,
                    'credit': adjustment
                })
            else:
                # Write-down: DR OCI, CR Investment
                entries.append({
                    'date': adj['date'],
                    'type': 'FMV Adjustment',
                    'asset_name': security,
                    'account_code': oci_account['code'],
                    'account_name': oci_account['name'],
                    'description': desc,
                    'debit': abs(adjustment),
                    'credit': 0.0
                })
                entries.append({
                    'date': adj['date'],
                    'type': 'FMV Adjustment',
                    'asset_name': security,
                    'account_code': inv_account['code'],
                    'account_name': inv_account['name'],
                    'description': desc,
                    'debit': 0.0,
                    'credit': abs(adjustment)
                })
        else:
            # Trading: Unrealized goes to P&L
            if adjustment > 0:
                gain_account = SECURITIES_ACCOUNTS['unrealized_gain']
                entries.append({
                    'date': adj['date'],
                    'type': 'FMV Adjustment',
                    'asset_name': security,
                    'account_code': inv_account['code'],
                    'account_name': inv_account['name'],
                    'description': desc,
                    'debit': adjustment,
                    'credit': 0.0
                })
                entries.append({
                    'date': adj['date'],
                    'type': 'FMV Adjustment',
                    'asset_name': security,
                    'account_code': gain_account['code'],
                    'account_name': gain_account['name'],
                    'description': desc,
                    'debit': 0.0,
                    'credit': adjustment
                })
            else:
                loss_account = SECURITIES_ACCOUNTS['unrealized_loss']
                entries.append({
                    'date': adj['date'],
                    'type': 'FMV Adjustment',
                    'asset_name': security,
                    'account_code': loss_account['code'],
                    'account_name': loss_account['name'],
                    'description': desc,
                    'debit': abs(adjustment),
                    'credit': 0.0
                })
                entries.append({
                    'date': adj['date'],
                    'type': 'FMV Adjustment',
                    'asset_name': security,
                    'account_code': inv_account['code'],
                    'account_name': inv_account['name'],
                    'description': desc,
                    'debit': 0.0,
                    'credit': abs(adjustment)
                })

    # Sort by date, then by type for consistent ordering
    entries.sort(key=lambda e: (e['date'], e['type'], e['asset_name']))
    return entries


def save_rollforward_xlsx(rollforward_data, filepath, report_date, period_begin):
    """Save rollforward to XLSX in standard format with color-coding.

    Color standards:
    - Blue: Hardcoded values (beginning balances, purchases, sales amounts)
    - Black: Formula cells (fair value calculations, totals)
    """
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not installed. XLSX output skipped.")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Securities Rollforward"

    # Styles
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    thin_border = Border(bottom=Side(style='thin'))
    double_border = Border(bottom=Side(style='double'))

    # Separate AFS and Trading
    afs_securities = {k: v for k, v in rollforward_data.items() if v['security_type'] == 'AFS'}
    trading_securities = {k: v for k, v in rollforward_data.items() if v['security_type'] == 'Trading'}

    row = 1

    # ============ AFS SECTION ============
    if afs_securities:
        # Title
        ws[f'A{row}'] = "AVAILABLE FOR SALE SECURITIES ROLLFORWARD"
        ws[f'A{row}'].font = title_font
        row += 1
        ws[f'A{row}'] = f"For the Period {period_begin} to {report_date}"
        row += 2

        # Column Headers
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=2, value="Cost Basis")
        ws.cell(row=row, column=3, value="Unrealized G/L")
        ws.cell(row=row, column=4, value="Fair Value")
        for col in range(2, 5):
            ws.cell(row=row, column=col).font = header_font
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        row += 1

        # Calculate totals
        total_beg_cost = sum(d['beginning_cost'] for d in afs_securities.values())
        total_beg_fmv = sum(d['beginning_fmv_adj'] for d in afs_securities.values())
        total_purchases = sum(d['purchases'] for d in afs_securities.values())
        total_sales = sum(d['sales_cost_basis'] for d in afs_securities.values())
        total_reclassification = sum(d['sales_fmv_adj_release'] for d in afs_securities.values())
        total_mtm = sum(d['period_fmv_adj'] for d in afs_securities.values())
        total_end_cost = sum(d['ending_cost'] for d in afs_securities.values())
        total_end_fmv = sum(d['ending_fmv_adj'] for d in afs_securities.values())

        # Beginning Balance - values are inputs (blue), FV is formula (black)
        beg_row = row
        ws.cell(row=row, column=1, value="Beginning Balance")
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=2, value=total_beg_cost)
        ws.cell(row=row, column=3, value=total_beg_fmv)
        ws.cell(row=row, column=4, value=f"=B{row}+C{row}")
        for col in range(2, 4):
            ws.cell(row=row, column=col).number_format = currency_format
            if STYLES_AVAILABLE:
                ws.cell(row=row, column=col).font = input_style()
        ws.cell(row=row, column=4).number_format = currency_format
        if STYLES_AVAILABLE:
            ws.cell(row=row, column=4).font = formula_style()
        row += 2

        # Purchases - cost is input (blue), FV is formula (black)
        purch_row = row
        ws.cell(row=row, column=1, value="Purchases")
        ws.cell(row=row, column=2, value=total_purchases)
        ws.cell(row=row, column=3, value="—")
        ws.cell(row=row, column=4, value=f"=B{row}")
        ws.cell(row=row, column=2).number_format = currency_format
        if STYLES_AVAILABLE:
            ws.cell(row=row, column=2).font = input_style()
        ws.cell(row=row, column=4).number_format = currency_format
        if STYLES_AVAILABLE:
            ws.cell(row=row, column=4).font = formula_style()
        row += 1

        # Sales (at cost) - cost is input (blue), FV is formula (black)
        sales_row = row
        ws.cell(row=row, column=1, value="Sales (at cost)")
        ws.cell(row=row, column=2, value=-total_sales if total_sales > 0 else 0)
        ws.cell(row=row, column=3, value="—")
        ws.cell(row=row, column=4, value=f"=B{row}")
        ws.cell(row=row, column=2).number_format = currency_format
        if STYLES_AVAILABLE:
            ws.cell(row=row, column=2).font = input_style()
        ws.cell(row=row, column=4).number_format = currency_format
        if STYLES_AVAILABLE:
            ws.cell(row=row, column=4).font = formula_style()
        row += 1

        # Reclassification to realized (OCI release) - unrealized is input (blue), FV is formula (black)
        reclass_row = row
        ws.cell(row=row, column=1, value="Reclassification to realized")
        ws.cell(row=row, column=2, value="—")
        ws.cell(row=row, column=3, value=-total_reclassification if total_reclassification > 0 else 0)
        ws.cell(row=row, column=4, value=f"=C{row}")
        ws.cell(row=row, column=3).number_format = currency_format
        if STYLES_AVAILABLE:
            ws.cell(row=row, column=3).font = input_style()
        ws.cell(row=row, column=4).number_format = currency_format
        if STYLES_AVAILABLE:
            ws.cell(row=row, column=4).font = formula_style()
        row += 1

        # Mark-to-market adjustment - adjustment is input (blue), FV is formula (black)
        mtm_row = row
        ws.cell(row=row, column=1, value="Mark-to-market adjustment")
        ws.cell(row=row, column=2, value="—")
        ws.cell(row=row, column=3, value=total_mtm)
        ws.cell(row=row, column=4, value=f"=C{row}")
        ws.cell(row=row, column=3).number_format = currency_format
        if STYLES_AVAILABLE:
            ws.cell(row=row, column=3).font = input_style()
        ws.cell(row=row, column=4).number_format = currency_format
        if STYLES_AVAILABLE:
            ws.cell(row=row, column=4).font = formula_style()
        row += 2

        # Ending Balance - all formulas (black with bold)
        end_row = row
        ws.cell(row=row, column=1, value="Ending Balance")
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=2, value=f"=B{beg_row}+B{purch_row}+B{sales_row}")
        ws.cell(row=row, column=3, value=f"=C{beg_row}+C{reclass_row}+C{mtm_row}")
        ws.cell(row=row, column=4, value=f"=B{row}+C{row}")
        for col in range(2, 5):
            ws.cell(row=row, column=col).number_format = currency_format
            ws.cell(row=row, column=col).border = double_border
            if STYLES_AVAILABLE:
                ws.cell(row=row, column=col).font = formula_style(bold=True)
            else:
                ws.cell(row=row, column=col).font = header_font
        row += 4

        # ============ REALIZED GAINS/LOSSES SCHEDULE ============
        ws.cell(row=row, column=1, value="REALIZED GAINS/LOSSES")
        ws.cell(row=row, column=1).font = title_font
        row += 1
        ws.cell(row=row, column=1, value=f"For the Period {period_begin} to {report_date}")
        row += 2

        # Headers
        ws.cell(row=row, column=1, value="Security")
        ws.cell(row=row, column=2, value="Proceeds")
        ws.cell(row=row, column=3, value="Cost Basis")
        ws.cell(row=row, column=4, value="Realized G/L")
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = header_font
            if col > 1:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        row += 1

        realized_start = row
        has_realized = False
        for security, data in sorted(afs_securities.items()):
            if data['realized_gain_loss'] != 0:
                has_realized = True
                ws.cell(row=row, column=1, value=security)
                ws.cell(row=row, column=2, value=data['sales_proceeds'])
                ws.cell(row=row, column=3, value=data['sales_cost_basis'])
                ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
                # Proceeds and cost basis are inputs (blue), G/L is formula (black)
                for col in range(2, 4):
                    ws.cell(row=row, column=col).number_format = currency_format
                    if STYLES_AVAILABLE:
                        ws.cell(row=row, column=col).font = input_style()
                ws.cell(row=row, column=4).number_format = currency_format
                if STYLES_AVAILABLE:
                    ws.cell(row=row, column=4).font = formula_style()
                row += 1

        if not has_realized:
            ws.cell(row=row, column=1, value="(No sales during period)")
            row += 1
        realized_end = row - 1

        # Total - all formulas (black with bold)
        if has_realized:
            row += 1
            ws.cell(row=row, column=1, value="Total Realized G/L")
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=2, value=f"=SUM(B{realized_start}:B{realized_end})")
            ws.cell(row=row, column=3, value=f"=SUM(C{realized_start}:C{realized_end})")
            ws.cell(row=row, column=4, value=f"=SUM(D{realized_start}:D{realized_end})")
            for col in range(2, 5):
                ws.cell(row=row, column=col).number_format = currency_format
                ws.cell(row=row, column=col).border = double_border
                if STYLES_AVAILABLE:
                    ws.cell(row=row, column=col).font = formula_style(bold=True)
                else:
                    ws.cell(row=row, column=col).font = header_font
        row += 4

        # ============ HOLDINGS DETAIL ============
        ws.cell(row=row, column=1, value="ENDING HOLDINGS DETAIL")
        ws.cell(row=row, column=1).font = title_font
        row += 1
        ws.cell(row=row, column=1, value=f"As of {report_date}")
        row += 2

        # Headers
        ws.cell(row=row, column=1, value="Security")
        ws.cell(row=row, column=2, value="Cost Basis")
        ws.cell(row=row, column=3, value="Unrealized G/L")
        ws.cell(row=row, column=4, value="Fair Value")
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = header_font
            if col > 1:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        row += 1

        detail_start = row
        for security, data in sorted(afs_securities.items()):
            if data['ending_cost'] > 0:
                ws.cell(row=row, column=1, value=security)
                ws.cell(row=row, column=2, value=data['ending_cost'])
                ws.cell(row=row, column=3, value=data['ending_fmv_adj'])
                ws.cell(row=row, column=4, value=f"=B{row}+C{row}")
                # Cost and unrealized are inputs (blue), FV is formula (black)
                for col in range(2, 4):
                    ws.cell(row=row, column=col).number_format = currency_format
                    if STYLES_AVAILABLE:
                        ws.cell(row=row, column=col).font = input_style()
                ws.cell(row=row, column=4).number_format = currency_format
                if STYLES_AVAILABLE:
                    ws.cell(row=row, column=4).font = formula_style()
                row += 1
        detail_end = row - 1

        # Total - all formulas (black with bold)
        row += 1
        ws.cell(row=row, column=1, value="Total")
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=2, value=f"=SUM(B{detail_start}:B{detail_end})")
        ws.cell(row=row, column=3, value=f"=SUM(C{detail_start}:C{detail_end})")
        ws.cell(row=row, column=4, value=f"=SUM(D{detail_start}:D{detail_end})")
        for col in range(2, 5):
            ws.cell(row=row, column=col).number_format = currency_format
            ws.cell(row=row, column=col).border = double_border
            if STYLES_AVAILABLE:
                ws.cell(row=row, column=col).font = formula_style(bold=True)
            else:
                ws.cell(row=row, column=col).font = header_font

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    wb.save(filepath)
    return True


def save_rollforward_csv(rollforward_data, filepath, report_date, period_begin):
    """Save rollforward to CSV in standard format."""
    import csv

    afs_securities = {k: v for k, v in rollforward_data.items() if v['security_type'] == 'AFS'}

    # Calculate totals
    total_beg_cost = sum(d['beginning_cost'] for d in afs_securities.values())
    total_beg_fmv = sum(d['beginning_fmv_adj'] for d in afs_securities.values())
    total_purchases = sum(d['purchases'] for d in afs_securities.values())
    total_sales = sum(d['sales_cost_basis'] for d in afs_securities.values())
    total_reclassification = sum(d['sales_fmv_adj_release'] for d in afs_securities.values())
    total_mtm = sum(d['period_fmv_adj'] for d in afs_securities.values())
    total_end_cost = sum(d['ending_cost'] for d in afs_securities.values())
    total_end_fmv = sum(d['ending_fmv_adj'] for d in afs_securities.values())

    with open(filepath, 'w', newline='') as f:
        writer = csv.writer(f)

        # Rollforward section
        writer.writerow(['AVAILABLE FOR SALE SECURITIES ROLLFORWARD'])
        writer.writerow([f'For the Period {period_begin} to {report_date}'])
        writer.writerow([])
        writer.writerow(['', 'Cost Basis', 'Unrealized G/L', 'Fair Value'])
        writer.writerow(['Beginning Balance', f'{total_beg_cost:.2f}', f'{total_beg_fmv:.2f}', f'{total_beg_cost + total_beg_fmv:.2f}'])
        writer.writerow(['Purchases', f'{total_purchases:.2f}', '—', f'{total_purchases:.2f}'])
        writer.writerow(['Sales (at cost)', f'{-total_sales:.2f}', '—', f'{-total_sales:.2f}'])
        writer.writerow(['Reclassification to realized', '—', f'{-total_reclassification:.2f}', f'{-total_reclassification:.2f}'])
        writer.writerow(['Mark-to-market adjustment', '—', f'{total_mtm:.2f}', f'{total_mtm:.2f}'])
        writer.writerow(['Ending Balance', f'{total_end_cost:.2f}', f'{total_end_fmv:.2f}', f'{total_end_cost + total_end_fmv:.2f}'])
        writer.writerow([])

        # Realized G/L section
        writer.writerow(['REALIZED GAINS/LOSSES'])
        writer.writerow([f'For the Period {period_begin} to {report_date}'])
        writer.writerow([])
        writer.writerow(['Security', 'Proceeds', 'Cost Basis', 'Realized G/L'])

        total_proceeds = 0
        total_cost_sold = 0
        total_realized = 0
        for security, data in sorted(afs_securities.items()):
            if data['realized_gain_loss'] != 0:
                writer.writerow([
                    security,
                    f'{data["sales_proceeds"]:.2f}',
                    f'{data["sales_cost_basis"]:.2f}',
                    f'{data["realized_gain_loss"]:.2f}'
                ])
                total_proceeds += data['sales_proceeds']
                total_cost_sold += data['sales_cost_basis']
                total_realized += data['realized_gain_loss']

        if total_realized != 0:
            writer.writerow([])
            writer.writerow(['Total', f'{total_proceeds:.2f}', f'{total_cost_sold:.2f}', f'{total_realized:.2f}'])
        writer.writerow([])

        # Holdings detail
        writer.writerow(['ENDING HOLDINGS DETAIL'])
        writer.writerow([f'As of {report_date}'])
        writer.writerow([])
        writer.writerow(['Security', 'Cost Basis', 'Unrealized G/L', 'Fair Value'])

        for security, data in sorted(afs_securities.items()):
            if data['ending_cost'] > 0:
                writer.writerow([
                    security,
                    f'{data["ending_cost"]:.2f}',
                    f'{data["ending_fmv_adj"]:.2f}',
                    f'{data["ending_cost"] + data["ending_fmv_adj"]:.2f}'
                ])

        writer.writerow([])
        writer.writerow(['Total', f'{total_end_cost:.2f}', f'{total_end_fmv:.2f}', f'{total_end_cost + total_end_fmv:.2f}'])


def main():
    args = parse_args()

    # Calculate period begin if using period type
    report_date = args.report_date
    if args.period_begin:
        period_begin = args.period_begin
        period_label = f"{period_begin} to {report_date}"
    else:
        period_begin = calculate_period_begin(report_date, args.period_type)
        period_label = get_period_label(period_begin, report_date, args.period_type)

    # Load register
    register = load_securities_register(args.output_dir)

    if not register.get('transactions') and not register.get('holdings'):
        print("No securities data found in register.")
        return

    # Calculate rollforward data
    rollforward_data = calculate_rollforward_data(register, report_date, period_begin)

    if not rollforward_data:
        print("No securities activity to report.")
        return

    # Create timestamped output folder
    run_timestamp = get_or_create_run_timestamp(args.run_timestamp)
    run_output_dir = os.path.join(args.output_dir, run_timestamp)
    os.makedirs(run_output_dir, exist_ok=True)

    csv_path = os.path.join(run_output_dir, f'securities_rollforward_{report_date}.csv')
    xlsx_path = os.path.join(run_output_dir, f'securities_rollforward_{report_date}.xlsx')

    save_rollforward_csv(rollforward_data, csv_path, report_date, period_begin)
    xlsx_ok = save_rollforward_xlsx(rollforward_data, xlsx_path, report_date, period_begin)

    # Generate consolidated journal entries
    journal_entries = generate_journal_entries(register, period_begin, report_date)
    journal_csv_path = os.path.join(run_output_dir, f'journal_{report_date}.csv')
    journal_xlsx_path = os.path.join(run_output_dir, f'journal_{report_date}.xlsx')

    journal_csv_ok = False
    journal_xlsx_ok = False
    if journal_entries:
        save_journal_csv(journal_entries, journal_csv_path)
        journal_csv_ok = True
        journal_xlsx_ok = save_journal_xlsx(journal_entries, journal_xlsx_path,
                                             title="Securities Journal")

    # Calculate totals for summary
    afs_securities = {k: v for k, v in rollforward_data.items() if v['security_type'] == 'AFS'}
    total_beg_cost = sum(d['beginning_cost'] for d in afs_securities.values())
    total_beg_fmv = sum(d['beginning_fmv_adj'] for d in afs_securities.values())
    total_purchases = sum(d['purchases'] for d in afs_securities.values())
    total_sales = sum(d['sales_cost_basis'] for d in afs_securities.values())
    total_proceeds = sum(d['sales_proceeds'] for d in afs_securities.values())
    total_realized = sum(d['realized_gain_loss'] for d in afs_securities.values())
    total_mtm = sum(d['period_fmv_adj'] for d in afs_securities.values())
    total_end_cost = sum(d['ending_cost'] for d in afs_securities.values())
    total_end_fmv = sum(d['ending_fmv_adj'] for d in afs_securities.values())

    # Print summary
    print(f"\n{'='*60}")
    print(f"SECURITIES ROLLFORWARD")
    print(f"{'='*60}")
    print(f"\nPeriod: {period_label}")

    print(f"\n--- Rollforward ---")
    print(f"{'':30} {'Cost Basis':>14} {'Unrealized':>14} {'Fair Value':>14}")
    print(f"{'Beginning Balance':30} ${total_beg_cost:>12,.2f} ${total_beg_fmv:>12,.2f} ${total_beg_cost + total_beg_fmv:>12,.2f}")
    print(f"{'Purchases':30} ${total_purchases:>12,.2f} {'—':>13} ${total_purchases:>12,.2f}")
    print(f"{'Sales (at cost)':30} ${-total_sales:>12,.2f} {'—':>13} ${-total_sales:>12,.2f}")
    print(f"{'Mark-to-market':30} {'—':>13} ${total_mtm:>12,.2f} ${total_mtm:>12,.2f}")
    print(f"{'Ending Balance':30} ${total_end_cost:>12,.2f} ${total_end_fmv:>12,.2f} ${total_end_cost + total_end_fmv:>12,.2f}")

    if total_realized != 0:
        print(f"\n--- Realized Gains/Losses ---")
        print(f"Proceeds: ${total_proceeds:,.2f}")
        print(f"Cost Basis: ${total_sales:,.2f}")
        if total_realized >= 0:
            print(f"REALIZED GAIN: ${total_realized:,.2f}")
        else:
            print(f"REALIZED LOSS: ${abs(total_realized):,.2f}")

    print(f"\n--- Output Files ---")
    print(f"Rollforward (CSV): {csv_path}")
    if xlsx_ok:
        print(f"Rollforward (XLSX): {xlsx_path}")
    if journal_csv_ok:
        print(f"Journal (CSV): {journal_csv_path}")
    if journal_xlsx_ok:
        print(f"Journal (XLSX): {journal_xlsx_path}")
    print(f"\n{'='*60}")


if __name__ == '__main__':
    main()
