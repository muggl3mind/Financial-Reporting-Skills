#!/usr/bin/env python3
"""
Securities Sale Script
Records the sale of investment securities with realized gain/loss calculation.
"""

import argparse
import json
import os
import sys
from datetime import datetime
import uuid

# Add shared module to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LIB_PATH = os.path.join(SCRIPT_DIR, '..', '..', '..', '..', 'lib')
sys.path.insert(0, LIB_PATH)

from utils import (save_journal_csv, sanitize_filename,
                   OPENPYXL_AVAILABLE, CURRENCY_FORMAT, validate_positive_amount,
                   validate_positive_int, validate_date, get_or_create_run_timestamp)
from accounts import SECURITIES_ACCOUNTS, get_securities_account

if OPENPYXL_AVAILABLE:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side

# Import xlsx styles for color-coding
try:
    from xlsx_styles import input_style, formula_style
    STYLES_AVAILABLE = True
except ImportError:
    STYLES_AVAILABLE = False

# Use shared account constants
CASH_ACCOUNT = SECURITIES_ACCOUNTS['cash']
REALIZED_GAIN = SECURITIES_ACCOUNTS['realized_gain']
REALIZED_LOSS = SECURITIES_ACCOUNTS['realized_loss']


def parse_args():
    parser = argparse.ArgumentParser(description='Record securities sale')
    parser.add_argument('--name', required=True, help='Security name')
    parser.add_argument('--date', required=True, help='Sale date (YYYY-MM-DD)')
    parser.add_argument('--shares', required=True, type=int, help='Number of shares sold')
    parser.add_argument('--cost-basis', required=True, type=float, help='Total cost basis of shares sold')
    parser.add_argument('--proceeds', required=True, type=float, help='Total sale proceeds')
    parser.add_argument('--security-type', default='AFS',
                        choices=['AFS', 'Trading'],
                        help='Security classification (default: AFS)')
    parser.add_argument('--output-dir', default='data/output/investments',
                        help='Output directory')
    parser.add_argument('--run-timestamp', default=None,
                        help='Timestamp for output folder (YYYY-MM-DD_HH-MM). If not provided, uses current time.')
    return parser.parse_args()


def generate_sale_journal(name, date, shares, cost_basis, proceeds, security_type):
    """Generate journal entries for securities sale."""
    gain_loss = proceeds - cost_basis
    entries = []
    description = f"Sale of {shares} shares of {name}"

    # Determine investment account based on security type
    inv_account = get_securities_account(security_type)

    # Debit: Cash (proceeds)
    entries.append({
        'date': date,
        'type': 'Securities Sale',
        'asset_name': name,
        'account_code': CASH_ACCOUNT['code'],
        'account_name': CASH_ACCOUNT['name'],
        'description': description,
        'debit': proceeds,
        'credit': 0.00
    })

    # If loss, debit Loss account
    if gain_loss < 0:
        entries.append({
            'date': date,
            'type': 'Securities Sale',
            'asset_name': name,
            'account_code': REALIZED_LOSS['code'],
            'account_name': REALIZED_LOSS['name'],
            'description': description,
            'debit': abs(gain_loss),
            'credit': 0.00
        })

    # Credit: Investment account (cost basis)
    entries.append({
        'date': date,
        'type': 'Securities Sale',
        'asset_name': name,
        'account_code': inv_account['code'],
        'account_name': inv_account['name'],
        'description': description,
        'debit': 0.00,
        'credit': cost_basis
    })

    # If gain, credit Gain account
    if gain_loss >= 0:
        entries.append({
            'date': date,
            'type': 'Securities Sale',
            'asset_name': name,
            'account_code': REALIZED_GAIN['code'],
            'account_name': REALIZED_GAIN['name'],
            'description': description,
            'debit': 0.00,
            'credit': gain_loss
        })

    return entries, {
        'type': 'sale',
        'security': name,
        'date': date,
        'shares': shares,
        'cost_basis': cost_basis,
        'proceeds': proceeds,
        'gain_loss': gain_loss,
        'is_gain': gain_loss >= 0,
        'security_type': security_type
    }


def update_securities_register(transaction_info, register_path):
    """Update securities register JSON file."""
    if os.path.exists(register_path):
        with open(register_path, 'r') as f:
            register = json.load(f)
    else:
        register = {'transactions': [], 'holdings': {}, 'metadata': {'total_transactions': 0}}

    # Add transaction
    transaction_id = f"txn_{str(uuid.uuid4())[:8]}"
    transaction_info['id'] = transaction_id
    transaction_info['recorded_at'] = datetime.now().isoformat()
    register['transactions'].append(transaction_info)

    # Update holdings
    security = transaction_info['security']
    if security in register['holdings']:
        register['holdings'][security]['shares'] -= transaction_info['shares']
        register['holdings'][security]['cost_basis'] -= transaction_info['cost_basis']

        # Clean up if no shares left
        if register['holdings'][security]['shares'] <= 0:
            del register['holdings'][security]

    register['metadata']['total_transactions'] = len(register['transactions'])
    register['metadata']['last_updated'] = datetime.now().isoformat()

    with open(register_path, 'w') as f:
        json.dump(register, f, indent=2)

    return transaction_id


def save_sale_xlsx(entries, filepath, transaction_info, holding_info=None):
    """Save sale journal to XLSX with gain/loss calculation summary.

    Includes:
    - Summary section with full gain/loss calculation
    - Cost basis derivation from original purchase
    - Journal entries section

    Args:
        entries: Journal entry list
        filepath: Output path
        transaction_info: Sale transaction details
        holding_info: Optional dict with original purchase info for cost basis calc
    """
    if not OPENPYXL_AVAILABLE:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Securities Sale"

    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    thin_border = Border(bottom=Side(style='thin'))
    double_border = Border(bottom=Side(style='double'))

    row = 1

    # === SUMMARY SECTION ===
    ws.cell(row=row, column=1, value="SECURITIES SALE - GAIN/LOSS CALCULATION")
    ws.cell(row=row, column=1).font = title_font
    row += 2

    # Security details
    ws.cell(row=row, column=1, value="Security:")
    ws.cell(row=row, column=2, value=transaction_info['security'])
    row += 1

    ws.cell(row=row, column=1, value="Sale Date:")
    ws.cell(row=row, column=2, value=transaction_info['date'])
    row += 1

    ws.cell(row=row, column=1, value="Classification:")
    ws.cell(row=row, column=2, value=transaction_info['security_type'])
    row += 2

    # === COST BASIS CALCULATION ===
    ws.cell(row=row, column=1, value="COST BASIS CALCULATION")
    ws.cell(row=row, column=1).font = header_font
    row += 1

    if holding_info and holding_info.get('total_shares_before_sale'):
        # Full calculation with formulas
        total_shares_row = row
        ws.cell(row=row, column=1, value="Total Shares (before sale):")
        cell = ws.cell(row=row, column=2, value=holding_info['total_shares_before_sale'])
        if STYLES_AVAILABLE:
            cell.font = input_style()
        row += 1

        total_cost_row = row
        ws.cell(row=row, column=1, value="Total Cost Basis:")
        cell = ws.cell(row=row, column=2, value=holding_info['total_cost_before_sale'])
        cell.number_format = CURRENCY_FORMAT
        if STYLES_AVAILABLE:
            cell.font = input_style()
        row += 1

        cost_per_share_row = row
        ws.cell(row=row, column=1, value="Cost Per Share:")
        cell = ws.cell(row=row, column=2, value=f"=B{total_cost_row}/B{total_shares_row}")
        cell.number_format = CURRENCY_FORMAT
        cell.border = thin_border
        if STYLES_AVAILABLE:
            cell.font = formula_style()
        row += 2

        shares_sold_row = row
        ws.cell(row=row, column=1, value="Shares Sold:")
        cell = ws.cell(row=row, column=2, value=transaction_info['shares'])
        if STYLES_AVAILABLE:
            cell.font = input_style()
        row += 1

        cost_basis_row = row
        ws.cell(row=row, column=1, value="Cost Basis of Shares Sold:")
        cell = ws.cell(row=row, column=2, value=f"=B{shares_sold_row}*B{cost_per_share_row}")
        cell.number_format = CURRENCY_FORMAT
        cell.border = thin_border
        if STYLES_AVAILABLE:
            cell.font = formula_style()
        row += 2
    else:
        # Fallback: just show the cost basis as input
        shares_sold_row = row
        ws.cell(row=row, column=1, value="Shares Sold:")
        cell = ws.cell(row=row, column=2, value=transaction_info['shares'])
        if STYLES_AVAILABLE:
            cell.font = input_style()
        row += 1

        cost_basis_row = row
        ws.cell(row=row, column=1, value="Cost Basis:")
        cell = ws.cell(row=row, column=2, value=transaction_info['cost_basis'])
        cell.number_format = CURRENCY_FORMAT
        if STYLES_AVAILABLE:
            cell.font = input_style()
        row += 2

    # === GAIN/LOSS CALCULATION ===
    ws.cell(row=row, column=1, value="GAIN/LOSS CALCULATION")
    ws.cell(row=row, column=1).font = header_font
    row += 1

    proceeds_row = row
    ws.cell(row=row, column=1, value="Sale Proceeds:")
    cell = ws.cell(row=row, column=2, value=transaction_info['proceeds'])
    cell.number_format = CURRENCY_FORMAT
    if STYLES_AVAILABLE:
        cell.font = input_style()
    row += 1

    ws.cell(row=row, column=1, value="Less: Cost Basis:")
    cell = ws.cell(row=row, column=2, value=f"=B{cost_basis_row}")
    cell.number_format = CURRENCY_FORMAT
    if STYLES_AVAILABLE:
        cell.font = formula_style()
    row += 1

    # Gain/Loss as formula
    if transaction_info['is_gain']:
        ws.cell(row=row, column=1, value="REALIZED GAIN:")
    else:
        ws.cell(row=row, column=1, value="REALIZED LOSS:")
    ws.cell(row=row, column=1).font = header_font

    cell = ws.cell(row=row, column=2, value=f"=B{proceeds_row}-B{cost_basis_row}")
    cell.number_format = CURRENCY_FORMAT
    cell.border = double_border
    if STYLES_AVAILABLE:
        cell.font = formula_style(bold=True)
    else:
        cell.font = header_font
    row += 3

    # === JOURNAL ENTRIES SECTION ===
    ws.cell(row=row, column=1, value="JOURNAL ENTRIES")
    ws.cell(row=row, column=1).font = title_font
    row += 2

    # Headers
    headers = ['Date', 'Account Code', 'Account Name', 'Description', 'Debit', 'Credit']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
    row += 1

    # Journal entries
    entry_start = row
    for entry in entries:
        ws.cell(row=row, column=1, value=entry['date'])
        ws.cell(row=row, column=2, value=entry['account_code'])
        ws.cell(row=row, column=3, value=entry['account_name'])
        ws.cell(row=row, column=4, value=entry['description'])

        if entry['debit'] > 0:
            cell = ws.cell(row=row, column=5, value=entry['debit'])
            cell.number_format = CURRENCY_FORMAT
            if STYLES_AVAILABLE:
                cell.font = input_style()

        if entry['credit'] > 0:
            cell = ws.cell(row=row, column=6, value=entry['credit'])
            cell.number_format = CURRENCY_FORMAT
            if STYLES_AVAILABLE:
                cell.font = input_style()

        row += 1

    entry_end = row - 1

    # Totals row
    row += 1
    ws.cell(row=row, column=4, value="TOTALS:").font = header_font

    debit_total = ws.cell(row=row, column=5, value=f"=SUM(E{entry_start}:E{entry_end})")
    debit_total.number_format = CURRENCY_FORMAT
    if STYLES_AVAILABLE:
        debit_total.font = formula_style(bold=True)
    else:
        debit_total.font = header_font

    credit_total = ws.cell(row=row, column=6, value=f"=SUM(F{entry_start}:F{entry_end})")
    credit_total.number_format = CURRENCY_FORMAT
    if STYLES_AVAILABLE:
        credit_total.font = formula_style(bold=True)
    else:
        credit_total.font = header_font

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    wb.save(filepath)
    return True


def main():
    args = parse_args()

    # Validate inputs
    try:
        validate_date(args.date, "Sale date")
        validate_positive_int(args.shares, "Shares")
        validate_positive_amount(args.cost_basis, "Cost basis")
        validate_positive_amount(args.proceeds, "Proceeds", allow_zero=True)
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)

    # Generate sale journal entry
    entries, transaction_info = generate_sale_journal(
        args.name, args.date, args.shares, args.cost_basis,
        args.proceeds, args.security_type
    )

    # Register stays at root level
    register_path = os.path.join(args.output_dir, 'securities_register.json')

    # Look up current holding info BEFORE updating register (for cost basis calculation)
    holding_info = None
    if os.path.exists(register_path):
        with open(register_path, 'r') as f:
            register = json.load(f)
        if args.name in register.get('holdings', {}):
            holding = register['holdings'][args.name]
            # Register has current holdings BEFORE this sale is recorded
            holding_info = {
                'total_shares_before_sale': holding['shares'],
                'total_cost_before_sale': holding['cost_basis']
            }

    # Create timestamped output folder
    run_timestamp = get_or_create_run_timestamp(args.run_timestamp)
    run_output_dir = os.path.join(args.output_dir, run_timestamp)
    os.makedirs(run_output_dir, exist_ok=True)

    safe_name = sanitize_filename(args.name)

    # Transaction files go to timestamped folder
    csv_path = os.path.join(run_output_dir, f'securities_sale_{safe_name}_{args.date}.csv')
    xlsx_path = os.path.join(run_output_dir, f'securities_sale_{safe_name}_{args.date}.xlsx')

    save_journal_csv(entries, csv_path)
    xlsx_ok = save_sale_xlsx(entries, xlsx_path, transaction_info, holding_info)
    txn_id = update_securities_register(transaction_info, register_path)

    # Print summary
    print(f"\n{'='*60}")
    print(f"SECURITIES SALE")
    print(f"{'='*60}")
    print(f"\nSecurity: {args.name}")
    print(f"Transaction ID: {txn_id}")
    print(f"Date: {args.date}")
    print(f"Shares Sold: {args.shares}")
    print(f"Cost Basis: ${args.cost_basis:,.2f}")
    print(f"Proceeds: ${args.proceeds:,.2f}")
    print(f"Classification: {args.security_type}")
    print(f"\n--- Realized Gain/Loss ---")
    if transaction_info['is_gain']:
        print(f"REALIZED GAIN: ${transaction_info['gain_loss']:,.2f}")
    else:
        print(f"REALIZED LOSS: ${abs(transaction_info['gain_loss']):,.2f}")
    print(f"\n--- Output Files ---")
    print(f"Journal (CSV): {csv_path}")
    if xlsx_ok:
        print(f"Journal (XLSX): {xlsx_path}")
    print(f"Securities Register: {register_path}")
    print(f"\n{'='*60}")


if __name__ == '__main__':
    main()
