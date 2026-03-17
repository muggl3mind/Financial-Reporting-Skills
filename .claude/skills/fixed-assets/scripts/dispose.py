#!/usr/bin/env python3
"""
Fixed Asset Disposal Script
Calculates depreciation through disposal date and generates disposal journal entries.
"""

import argparse
import json
import csv
import os
import sys
from datetime import datetime
import uuid

# Add shared and local modules to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LIB_PATH = os.path.join(SCRIPT_DIR, '..', '..', '..', '..', 'lib')
sys.path.insert(0, LIB_PATH)
sys.path.insert(0, SCRIPT_DIR)

from accounts import get_account_mapping, SECURITIES_ACCOUNTS, get_default_useful_life
from utils import (sanitize_filename, validate_positive_amount, validate_positive_int,
                   validate_date, get_or_create_run_timestamp)
from depreciation import (
    calculate_straight_line,
    generate_schedule,
    OPENPYXL_AVAILABLE
)

if OPENPYXL_AVAILABLE:
    from openpyxl import Workbook
    from openpyxl.styles import Font


def parse_args():
    parser = argparse.ArgumentParser(description='Generate fixed asset disposal journal entries')
    parser.add_argument('--name', required=True, help='Asset name')
    parser.add_argument('--category', required=True,
                        choices=['Equipment', 'Vehicle', 'Furniture', 'Computer',
                                 'Building', 'Software', 'Patent'],
                        help='Asset category')
    parser.add_argument('--purchase-date', required=True, help='Purchase date (YYYY-MM-DD)')
    parser.add_argument('--cost', required=True, type=float, help='Original cost')
    parser.add_argument('--salvage', required=True, type=float, help='Salvage value')
    parser.add_argument('--useful-life', type=int, default=None,
                        help='Useful life in years (optional, defaults to MACRS based on category)')
    parser.add_argument('--disposal-date', required=True, help='Disposal date (YYYY-MM-DD)')
    parser.add_argument('--proceeds', required=True, type=float, help='Sale proceeds')
    parser.add_argument('--output-dir', default='data/output/fixed-assets',
                        help='Output directory')
    parser.add_argument('--run-timestamp', default=None,
                        help='Timestamp for output folder (YYYY-MM-DD_HH-MM). If not provided, uses current time.')
    return parser.parse_args()


def calculate_accumulated_depreciation(schedule, disposal_date):
    """Calculate accumulated depreciation through the disposal date."""
    disposal_dt = datetime.strptime(disposal_date, '%Y-%m-%d')
    accumulated = 0.0

    for period in schedule:
        period_end = datetime.strptime(period['end_date'], '%Y-%m-%d')
        period_start = datetime.strptime(period['start_date'], '%Y-%m-%d')

        if period_end <= disposal_dt:
            # Full period before disposal
            accumulated += period['depreciation']
        elif period_start <= disposal_dt < period_end:
            # Partial period - prorate depreciation to disposal date
            total_days = (period_end - period_start).days + 1
            days_to_disposal = (disposal_dt - period_start).days + 1
            prorated_depr = period['depreciation'] * (days_to_disposal / total_days)
            accumulated += round(prorated_depr, 2)
            break
        else:
            # Period is after disposal date
            break

    return round(accumulated, 2)


def calculate_disposal(cost, accumulated_depreciation, proceeds):
    """Calculate gain or loss on disposal."""
    book_value = cost - accumulated_depreciation
    gain_loss = proceeds - book_value

    return {
        'cost': cost,
        'accumulated_depreciation': accumulated_depreciation,
        'book_value': round(book_value, 2),
        'proceeds': proceeds,
        'gain_loss': round(gain_loss, 2),
        'is_gain': gain_loss >= 0
    }


def generate_monthly_breakdown(purchase_date, disposal_date):
    """Generate monthly breakdown for pro-rata depreciation calculation.

    Returns list of dicts with:
    - period_name: e.g., "Mar 2025"
    - days: number of days in service this month
    - days_in_month: total days in the month
    """
    from calendar import monthrange

    purchase_dt = datetime.strptime(purchase_date, '%Y-%m-%d')
    disposal_dt = datetime.strptime(disposal_date, '%Y-%m-%d')

    breakdown = []
    current_year = purchase_dt.year
    current_month = purchase_dt.month

    while True:
        # Get days in this month
        _, days_in_month = monthrange(current_year, current_month)

        # Calculate month start and end
        month_start = datetime(current_year, current_month, 1)
        month_end = datetime(current_year, current_month, days_in_month)

        # Determine days in service for this month
        if current_year == purchase_dt.year and current_month == purchase_dt.month:
            # First month - start from purchase date
            start_day = purchase_dt.day
        else:
            start_day = 1

        if current_year == disposal_dt.year and current_month == disposal_dt.month:
            # Last month - end at disposal date
            end_day = disposal_dt.day
        else:
            end_day = days_in_month

        days = end_day - start_day + 1

        period_name = datetime(current_year, current_month, 1).strftime("%b %Y")

        breakdown.append({
            'period_name': period_name,
            'days': days,
            'days_in_month': days_in_month
        })

        # Check if we've reached the disposal month
        if current_year == disposal_dt.year and current_month == disposal_dt.month:
            break

        # Move to next month
        current_month += 1
        if current_month > 12:
            current_month = 1
            current_year += 1

        # Safety check - don't go past disposal date
        if datetime(current_year, current_month, 1) > disposal_dt:
            break

    return breakdown


def generate_disposal_journal(asset_name, disposal_info, accounts, disposal_date):
    """Generate journal entries for disposal."""
    entries = []
    description = f"Disposal of {asset_name}"

    # Entry 1: Debit Cash (proceeds)
    cash = SECURITIES_ACCOUNTS['cash']
    entries.append({
        'date': disposal_date,
        'account_code': cash['code'],
        'account_name': cash['name'],
        'description': description,
        'debit': disposal_info['proceeds'],
        'credit': 0.00
    })

    # Entry 2: Debit Accumulated Depreciation (remove contra-asset)
    entries.append({
        'date': disposal_date,
        'account_code': accounts['accum_code'],
        'account_name': accounts['accum_name'],
        'description': description,
        'debit': disposal_info['accumulated_depreciation'],
        'credit': 0.00
    })

    # Entry 3: Credit Asset account (remove asset at original cost)
    entries.append({
        'date': disposal_date,
        'account_code': accounts['asset_code'],
        'account_name': accounts['asset_name'],
        'description': description,
        'debit': 0.00,
        'credit': disposal_info['cost']
    })

    # Entry 4: Gain or Loss
    if disposal_info['is_gain']:
        entries.append({
            'date': disposal_date,
            'account_code': accounts['gain_code'],
            'account_name': accounts['gain_name'],
            'description': description,
            'debit': 0.00,
            'credit': abs(disposal_info['gain_loss'])
        })
    else:
        entries.append({
            'date': disposal_date,
            'account_code': accounts['loss_code'],
            'account_name': accounts['loss_name'],
            'description': description,
            'debit': abs(disposal_info['gain_loss']),
            'credit': 0.00
        })

    return entries


def save_disposal_journal_csv(entries, filepath):
    """Save disposal journal entries to CSV."""
    fieldnames = ['Date', 'Account Code', 'Account Name', 'Description', 'Debit', 'Credit']

    with open(filepath, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(fieldnames)
        for entry in entries:
            writer.writerow([
                entry['date'],
                entry['account_code'],
                entry['account_name'],
                entry['description'],
                f"{entry['debit']:.2f}" if entry['debit'] > 0 else '',
                f"{entry['credit']:.2f}" if entry['credit'] > 0 else ''
            ])


def save_disposal_journal_xlsx(entries, filepath, disposal_info, asset_name, calc_details, monthly_breakdown):
    """Save disposal journal entries to XLSX with monthly pro-rata formulas."""
    if not OPENPYXL_AVAILABLE:
        return False

    from openpyxl.styles import Border, Side, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Disposal Journal"

    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    thin_border = Border(bottom=Side(style='thin'))
    double_border = Border(bottom=Side(style='double'))

    # Title
    ws['A1'] = f"DISPOSAL OF FIXED ASSET: {asset_name}"
    ws['A1'].font = title_font

    # === ASSET INFORMATION ===
    row = 3
    ws[f'A{row}'] = "ASSET INFORMATION"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    # Original Cost
    ws[f'A{row}'] = "Original Cost:"
    ws[f'B{row}'] = disposal_info['cost']
    ws[f'B{row}'].number_format = currency_format
    cost_row = row
    row += 1

    # Salvage Value
    ws[f'A{row}'] = "Salvage Value:"
    ws[f'B{row}'] = calc_details['salvage']
    ws[f'B{row}'].number_format = currency_format
    salvage_row = row
    row += 1

    # Depreciable Amount (formula)
    ws[f'A{row}'] = "Depreciable Amount:"
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'] = f"=B{cost_row}-B{salvage_row}"
    ws[f'B{row}'].number_format = currency_format
    depr_amount_row = row
    row += 1

    # Useful Life
    ws[f'A{row}'] = "Useful Life (years):"
    ws[f'B{row}'] = calc_details['useful_life']
    life_row = row
    row += 1

    # Monthly Depreciation (formula)
    ws[f'A{row}'] = "Monthly Depreciation:"
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'] = f"=B{depr_amount_row}/B{life_row}/12"
    ws[f'B{row}'].number_format = currency_format
    monthly_depr_row = row
    row += 2

    # === ACCUMULATED DEPRECIATION BREAKDOWN ===
    ws[f'A{row}'] = "ACCUMULATED DEPRECIATION BREAKDOWN"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    # Headers for breakdown table
    ws[f'A{row}'] = "Period"
    ws[f'B{row}'] = "Days"
    ws[f'C{row}'] = "Days in Month"
    ws[f'D{row}'] = "Monthly Depr"
    ws[f'E{row}'] = "Pro-Rata Depr"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].alignment = Alignment(horizontal='right' if col != 'A' else 'left')
    row += 1

    breakdown_start = row
    for period in monthly_breakdown:
        ws[f'A{row}'] = period['period_name']
        ws[f'B{row}'] = period['days']
        ws[f'C{row}'] = period['days_in_month']
        ws[f'D{row}'] = f"=$B${monthly_depr_row}"
        ws[f'D{row}'].number_format = currency_format
        # Pro-rata formula: Monthly Depr * Days / Days in Month
        ws[f'E{row}'] = f"=ROUND(D{row}*B{row}/C{row},2)"
        ws[f'E{row}'].number_format = currency_format
        row += 1
    breakdown_end = row - 1

    # Total Accumulated Depreciation
    ws[f'A{row}'] = "TOTAL ACCUM DEPRECIATION:"
    ws[f'A{row}'].font = header_font
    ws[f'E{row}'] = f"=SUM(E{breakdown_start}:E{breakdown_end})"
    ws[f'E{row}'].number_format = currency_format
    ws[f'E{row}'].font = header_font
    ws[f'E{row}'].border = double_border
    accum_depr_row = row
    row += 2

    # === GAIN/LOSS CALCULATION ===
    ws[f'A{row}'] = "GAIN/LOSS CALCULATION"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    # Book Value at Disposal (formula)
    ws[f'A{row}'] = "Book Value at Disposal:"
    ws[f'B{row}'] = f"=B{cost_row}-E{accum_depr_row}"
    ws[f'B{row}'].number_format = currency_format
    book_value_row = row
    row += 1

    # Proceeds
    ws[f'A{row}'] = "Proceeds:"
    ws[f'B{row}'] = disposal_info['proceeds']
    ws[f'B{row}'].number_format = currency_format
    proceeds_row = row
    row += 1

    # Gain (Loss) on Disposal (formula)
    ws[f'A{row}'] = "Gain (Loss) on Disposal:"
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'] = f"=B{proceeds_row}-B{book_value_row}"
    ws[f'B{row}'].number_format = currency_format
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].border = double_border
    gain_loss_row = row
    row += 2

    # === JOURNAL ENTRY ===
    ws[f'A{row}'] = "JOURNAL ENTRY"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    headers = ['Date', 'Account Code', 'Account Name', 'Description', 'Debit', 'Credit']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
    row += 1

    journal_start = row
    for i, entry in enumerate(entries):
        ws.cell(row=row, column=1, value=entry['date'])
        ws.cell(row=row, column=2, value=entry['account_code'])
        ws.cell(row=row, column=3, value=entry['account_name'])
        ws.cell(row=row, column=4, value=entry['description'])

        # Use formula references for Accum Depr and Gain/Loss
        if entry['account_code'] in ['1600', '1650']:  # Accumulated Depreciation
            ws.cell(row=row, column=5, value=f"=E{accum_depr_row}")
            ws.cell(row=row, column=5).number_format = currency_format
        elif entry['account_code'] in ['7200', '8100']:  # Gain or Loss
            if disposal_info['is_gain']:
                ws.cell(row=row, column=6, value=f"=B{gain_loss_row}")
            else:
                ws.cell(row=row, column=5, value=f"=ABS(B{gain_loss_row})")
                ws.cell(row=row, column=5).number_format = currency_format
            ws.cell(row=row, column=6).number_format = currency_format
        elif entry['debit'] > 0:
            ws.cell(row=row, column=5, value=entry['debit'])
            ws.cell(row=row, column=5).number_format = currency_format
        elif entry['credit'] > 0:
            ws.cell(row=row, column=6, value=entry['credit'])
            ws.cell(row=row, column=6).number_format = currency_format
        row += 1
    journal_end = row - 1

    # Totals
    ws.cell(row=row, column=4, value="TOTALS:").font = header_font
    ws.cell(row=row, column=5, value=f"=SUM(E{journal_start}:E{journal_end})")
    ws.cell(row=row, column=5).number_format = currency_format
    ws.cell(row=row, column=5).font = header_font
    ws.cell(row=row, column=5).border = double_border
    ws.cell(row=row, column=6, value=f"=SUM(F{journal_start}:F{journal_end})")
    ws.cell(row=row, column=6).number_format = currency_format
    ws.cell(row=row, column=6).font = header_font
    ws.cell(row=row, column=6).border = double_border

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16

    wb.save(filepath)
    return True


def update_asset_register_disposal(output_dir, asset_name, disposal_date, disposal_info):
    """Update asset register with disposal information."""
    register_path = os.path.join(output_dir, 'asset_register.json')

    if os.path.exists(register_path):
        with open(register_path, 'r') as f:
            register = json.load(f)
    else:
        register = {'assets': [], 'metadata': {}}

    # Find the asset and update with disposal info
    for asset in register.get('assets', []):
        if asset['name'] == asset_name:
            asset['disposal_date'] = disposal_date
            asset['disposal_info'] = {
                'proceeds': disposal_info['proceeds'],
                'accumulated_depreciation': disposal_info['accumulated_depreciation'],
                'book_value': disposal_info['book_value'],
                'gain_loss': disposal_info['gain_loss'],
                'is_gain': disposal_info['is_gain']
            }
            asset['status'] = 'disposed'
            break

    register['metadata']['last_updated'] = datetime.now().isoformat()

    with open(register_path, 'w') as f:
        json.dump(register, f, indent=2)


def main():
    args = parse_args()

    # Use provided useful_life or default based on category
    useful_life = args.useful_life if args.useful_life is not None else get_default_useful_life(args.category)

    # Validate inputs
    try:
        validate_date(args.purchase_date, "Purchase date")
        validate_date(args.disposal_date, "Disposal date")
        validate_positive_amount(args.cost, "Cost")
        validate_positive_amount(args.salvage, "Salvage value", allow_zero=True)
        validate_positive_int(useful_life, "Useful life")
        validate_positive_amount(args.proceeds, "Proceeds", allow_zero=True)
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)

    # Get account mapping (includes gain/loss accounts)
    accounts = get_account_mapping(args.category)

    # Calculate depreciation schedule
    calc_results = calculate_straight_line(
        args.cost, args.salvage, useful_life, 'monthly'
    )

    schedule = generate_schedule(
        args.name, args.purchase_date, args.cost, args.salvage,
        useful_life, 'monthly', calc_results, args.disposal_date
    )

    # Calculate accumulated depreciation through disposal date
    accumulated_depr = calculate_accumulated_depreciation(schedule, args.disposal_date)

    # Calculate days in service for formula display
    purchase_dt = datetime.strptime(args.purchase_date, '%Y-%m-%d')
    disposal_dt = datetime.strptime(args.disposal_date, '%Y-%m-%d')
    days_in_service = (disposal_dt - purchase_dt).days + 1

    # Calculate disposal gain/loss
    disposal_info = calculate_disposal(args.cost, accumulated_depr, args.proceeds)

    # Calculation details for formulas
    calc_details = {
        'salvage': args.salvage,
        'useful_life': useful_life,
        'days_in_service': days_in_service
    }

    # Generate monthly breakdown for XLSX
    monthly_breakdown = generate_monthly_breakdown(args.purchase_date, args.disposal_date)

    # Generate disposal journal entries
    journal_entries = generate_disposal_journal(
        args.name, disposal_info, accounts, args.disposal_date
    )

    # Create timestamped output folder
    run_timestamp = get_or_create_run_timestamp(args.run_timestamp)
    run_output_dir = os.path.join(args.output_dir, run_timestamp)
    os.makedirs(run_output_dir, exist_ok=True)

    safe_name = sanitize_filename(args.name)

    # Transaction files go to timestamped folder
    csv_path = os.path.join(run_output_dir, f'disposal_{safe_name}.csv')
    xlsx_path = os.path.join(run_output_dir, f'disposal_{safe_name}.xlsx')

    save_disposal_journal_csv(journal_entries, csv_path)
    xlsx_ok = save_disposal_journal_xlsx(journal_entries, xlsx_path, disposal_info, args.name, calc_details, monthly_breakdown)

    # Update asset register (stays at root level)
    update_asset_register_disposal(args.output_dir, args.name, args.disposal_date, disposal_info)

    # Print summary
    print(f"\n{'='*60}")
    print(f"FIXED ASSET DISPOSAL")
    print(f"{'='*60}")
    print(f"\nAsset: {args.name}")
    print(f"Category: {args.category}")
    print(f"Purchase Date: {args.purchase_date}")
    print(f"Disposal Date: {args.disposal_date}")
    print(f"\n--- Disposal Calculation ---")
    print(f"Original Cost: ${args.cost:,.2f}")
    print(f"Accumulated Depreciation: ${accumulated_depr:,.2f}")
    print(f"Book Value at Disposal: ${disposal_info['book_value']:,.2f}")
    print(f"Sale Proceeds: ${args.proceeds:,.2f}")

    if disposal_info['is_gain']:
        print(f"GAIN on Disposal: ${disposal_info['gain_loss']:,.2f}")
    else:
        print(f"LOSS on Disposal: ${abs(disposal_info['gain_loss']):,.2f}")

    print(f"\n--- Account Mapping ---")
    print(f"Asset Account: {accounts['asset_code']} - {accounts['asset_name']}")
    print(f"Accum. Account: {accounts['accum_code']} - {accounts['accum_name']}")
    if disposal_info['is_gain']:
        print(f"Gain Account: {accounts['gain_code']} - {accounts['gain_name']}")
    else:
        print(f"Loss Account: {accounts['loss_code']} - {accounts['loss_name']}")

    print(f"\n--- Output Files ---")
    print(f"Journal (CSV): {csv_path}")
    if xlsx_ok:
        print(f"Journal (XLSX): {xlsx_path}")
    print(f"\n{'='*60}")


if __name__ == '__main__':
    main()
