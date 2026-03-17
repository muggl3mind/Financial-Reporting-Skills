#!/usr/bin/env python3
"""
Depreciation Calculation Module
Core depreciation functions for fixed assets.
"""

import csv
import os
import sys
import json
from datetime import datetime, timedelta
import calendar
import uuid

# Add shared module to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LIB_PATH = os.path.join(SCRIPT_DIR, '..', '..', '..', '..', 'lib')
sys.path.insert(0, LIB_PATH)

from utils import sanitize_filename

# Try to import openpyxl for Excel output
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import xlsx styles for color-coding
try:
    from xlsx_styles import input_style, formula_style, style_value_cell, is_formula_value
    STYLES_AVAILABLE = True
except ImportError:
    STYLES_AVAILABLE = False


def add_months(date, months):
    """Add months to a date, handling month-end edge cases."""
    month = date.month - 1 + months
    year = date.year + month // 12
    month = month % 12 + 1
    day = min(date.day, calendar.monthrange(year, month)[1])
    return datetime(year, month, day)


def calculate_straight_line(cost, salvage, useful_life_years, period_type=None):
    """Calculate straight-line depreciation amounts.

    Always calculates monthly depreciation. Period type is accepted but ignored
    for backward compatibility - schedule is always monthly.
    """
    depreciable_amount = cost - salvage
    annual_depreciation = depreciable_amount / useful_life_years
    monthly_depreciation = annual_depreciation / 12
    total_periods = useful_life_years * 12  # Always monthly

    return {
        'depreciable_amount': round(depreciable_amount, 2),
        'annual_depreciation': round(annual_depreciation, 2),
        'monthly_depreciation': round(monthly_depreciation, 2),
        'total_periods': total_periods
    }


def get_period_end_date(start_date, period_type):
    """Calculate the end date for a given period."""
    if period_type == 'monthly':
        end_date = add_months(start_date, 1) - timedelta(days=1)
    elif period_type == 'quarterly':
        end_date = add_months(start_date, 3) - timedelta(days=1)
    else:  # annual
        end_date = add_months(start_date, 12) - timedelta(days=1)
    return end_date


def get_next_period_start(current_start, period_type):
    """Get the start date of the next period."""
    if period_type == 'monthly':
        return add_months(current_start, 1)
    elif period_type == 'quarterly':
        return add_months(current_start, 3)
    else:  # annual
        return add_months(current_start, 12)


def get_calendar_period_end(date, period_type):
    """Get the end of the calendar period containing the given date."""
    if period_type == 'monthly':
        # End of current month
        last_day = calendar.monthrange(date.year, date.month)[1]
        return datetime(date.year, date.month, last_day)
    elif period_type == 'quarterly':
        # End of current quarter
        quarter_end_months = {1: 3, 2: 3, 3: 3, 4: 6, 5: 6, 6: 6,
                             7: 9, 8: 9, 9: 9, 10: 12, 11: 12, 12: 12}
        end_month = quarter_end_months[date.month]
        last_day = calendar.monthrange(date.year, end_month)[1]
        return datetime(date.year, end_month, last_day)
    else:  # annual
        # End of current year
        return datetime(date.year, 12, 31)


def get_next_calendar_period_start(date, period_type):
    """Get the start of the next calendar period."""
    if period_type == 'monthly':
        if date.month == 12:
            return datetime(date.year + 1, 1, 1)
        return datetime(date.year, date.month + 1, 1)
    elif period_type == 'quarterly':
        quarter_start_months = {1: 4, 2: 4, 3: 4, 4: 7, 5: 7, 6: 7,
                               7: 10, 8: 10, 9: 10, 10: 1, 11: 1, 12: 1}
        next_month = quarter_start_months[date.month]
        next_year = date.year + 1 if date.month >= 10 else date.year
        return datetime(next_year, next_month, 1)
    else:  # annual
        return datetime(date.year + 1, 1, 1)


def get_full_period_days(period_type):
    """Get approximate days in a full period for proration."""
    if period_type == 'monthly':
        return 30  # Average month
    elif period_type == 'quarterly':
        return 91  # Average quarter
    else:  # annual
        return 365


def generate_schedule(asset_name, purchase_date, cost, salvage, useful_life,
                      period_type, calc_results, disposal_date=None):
    """Generate monthly depreciation schedule with day-based proration.

    Formula: Monthly Depreciation × (Days in Period / Days in Month)

    This formula can be exactly replicated in Excel:
        =ROUND(Monthly_Depr * Days / DAY(EOMONTH(Start_Date, 0)), 2)

    Args:
        asset_name: Name of the asset
        purchase_date: Purchase date string (YYYY-MM-DD)
        cost: Original cost of asset
        salvage: Salvage value
        useful_life: Useful life in years
        period_type: Ignored - schedule is always monthly
        calc_results: Dictionary from calculate_straight_line()
        disposal_date: Optional disposal date string (YYYY-MM-DD)

    Returns:
        List of schedule periods (monthly)
    """
    schedule = []

    purchase_dt = datetime.strptime(purchase_date, '%Y-%m-%d')
    disposal_dt = datetime.strptime(disposal_date, '%Y-%m-%d') if disposal_date else None
    book_value = cost
    accumulated = 0.0
    monthly_depr = calc_results['monthly_depreciation']

    period = 0
    current_start = purchase_dt

    # Continue until fully depreciated to salvage value (or hit disposal date)
    while book_value > salvage + 0.01:  # Small tolerance for rounding
        period += 1

        # Period end is always end of month
        period_end = get_calendar_period_end(current_start, 'monthly')

        # Check if disposal date falls in this period
        is_disposal_period = False
        if disposal_dt:
            if current_start <= disposal_dt <= period_end:
                period_end = disposal_dt
                is_disposal_period = True
            elif disposal_dt < current_start:
                break

        # Calculate days in this period and days in the month
        days_in_period = (period_end - current_start).days + 1
        days_in_month = calendar.monthrange(current_start.year, current_start.month)[1]

        # Proration formula: Monthly Depr × (Days in Period / Days in Month)
        period_depr = round(monthly_depr * days_in_period / days_in_month, 2)

        # Don't depreciate below salvage value
        if book_value - period_depr < salvage:
            period_depr = round(book_value - salvage, 2)

        beginning_book_value = book_value
        accumulated += period_depr
        book_value -= period_depr

        schedule.append({
            'period': period,
            'start_date': current_start.strftime('%Y-%m-%d'),
            'end_date': period_end.strftime('%Y-%m-%d'),
            'days': days_in_period,
            'days_in_month': days_in_month,
            'beginning_book_value': round(beginning_book_value, 2),
            'depreciation': round(period_depr, 2),
            'accumulated_depreciation': round(accumulated, 2),
            'ending_book_value': round(book_value, 2)
        })

        if is_disposal_period:
            break

        # Move to next month
        current_start = get_next_calendar_period_start(period_end, 'monthly')

        # Safety check
        if period > calc_results['total_periods'] + 12:
            break

    return schedule


def generate_journal_entries(asset_name, schedule, accounts, period_type):
    """Generate journal entries from the schedule."""
    entries = []

    period_label = {
        'monthly': 'Monthly',
        'quarterly': 'Quarterly',
        'annual': 'Annual'
    }

    for row in schedule:
        entry_id = row['period']
        date = row['end_date']
        description = f"{period_label[period_type]} depreciation - {asset_name}"
        amount = row['depreciation']

        # Debit: Expense account
        entries.append({
            'entry_id': entry_id,
            'date': date,
            'account_code': accounts['expense_code'],
            'account_name': accounts['expense_name'],
            'description': description,
            'debit': amount,
            'credit': 0.00
        })

        # Credit: Accumulated Depreciation
        entries.append({
            'entry_id': entry_id,
            'date': date,
            'account_code': accounts['accum_code'],
            'account_name': accounts['accum_name'],
            'description': description,
            'debit': 0.00,
            'credit': amount
        })

    return entries


def save_schedule_csv(schedule, filepath):
    """Save schedule to CSV file."""
    fieldnames = ['Period', 'Start Date', 'End Date', 'Days', 'Days in Month',
                  'Beginning Book Value', 'Depreciation', 'Accumulated Depreciation', 'Ending Book Value']

    with open(filepath, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(fieldnames)
        for row in schedule:
            writer.writerow([
                row['period'],
                row['start_date'],
                row['end_date'],
                row['days'],
                row['days_in_month'],
                f"{row['beginning_book_value']:.2f}",
                f"{row['depreciation']:.2f}",
                f"{row['accumulated_depreciation']:.2f}",
                f"{row['ending_book_value']:.2f}"
            ])


def save_journal_csv(entries, filepath):
    """Save journal entries to CSV file."""
    fieldnames = ['Entry ID', 'Date', 'Account Code', 'Account Name',
                  'Description', 'Debit', 'Credit']

    with open(filepath, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(fieldnames)
        for entry in entries:
            writer.writerow([
                entry['entry_id'],
                entry['date'],
                entry['account_code'],
                entry['account_name'],
                entry['description'],
                f"{entry['debit']:.2f}",
                f"{entry['credit']:.2f}"
            ])


def save_schedule_xlsx(schedule, filepath, asset_info):
    """Save schedule to XLSX file with Excel formulas and color-coding.

    Color standards:
    - Blue: Hardcoded inputs (cost, salvage, useful life)
    - Black: Formula cells (calculations)
    """
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not installed. XLSX output skipped.")
        print("Install with: pip install openpyxl")
        return False

    wb = Workbook()

    # --- Summary Sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Styles
    header_font = Font(bold=True)
    currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    # Asset Information
    ws_summary['A1'] = "DEPRECIATION SUMMARY"
    ws_summary['A1'].font = Font(bold=True, size=14)

    # Store useful_life as a number for formula reference
    useful_life = asset_info['useful_life']

    # Define which rows contain inputs vs formulas for color-coding
    # Rows 6, 7, 8 are inputs (cost, salvage, useful life)
    # Rows 11, 12, 13 are formulas (calculations)
    input_labels = ('Cost:', 'Salvage Value:', 'Useful Life (years):')
    formula_labels = ('Depreciable Amount:', 'Annual Depreciation:', 'Monthly Depreciation:')

    summary_data = [
        ('Asset Name:', asset_info['name']),           # Row 3
        ('Category:', asset_info['category']),          # Row 4
        ('Purchase Date:', asset_info['purchase_date']),# Row 5
        ('Cost:', asset_info['cost']),                  # Row 6  - D6
        ('Salvage Value:', asset_info['salvage']),      # Row 7  - D7
        ('Useful Life (years):', useful_life),          # Row 8  - D8
        ('Method:', 'Straight-Line'),                   # Row 9
        ('', ''),                                        # Row 10
        ('Depreciable Amount:', '=D6-D7'),              # Row 11 - D11 = Cost minus Salvage
        ('Annual Depreciation:', '=D11/D8'),            # Row 12 - D12 = Depreciable / Years
        ('Monthly Depreciation:', '=D12/12'),           # Row 13 - D13 = Annual / 12
    ]

    currency_labels = ('Cost:', 'Salvage Value:', 'Depreciable Amount:',
                       'Annual Depreciation:', 'Monthly Depreciation:')

    for i, (label, value) in enumerate(summary_data, start=3):
        ws_summary[f'C{i}'] = label
        ws_summary[f'C{i}'].font = header_font
        ws_summary[f'D{i}'] = value
        if label in currency_labels:
            ws_summary[f'D{i}'].number_format = currency_format

        # Apply color-coding based on cell type
        if STYLES_AVAILABLE:
            if label in input_labels:
                ws_summary[f'D{i}'].font = input_style()
            elif label in formula_labels:
                ws_summary[f'D{i}'].font = formula_style()

    # --- Schedule Sheet ---
    ws_schedule = wb.create_sheet("Schedule")

    # Headers - includes Days in Month for formula transparency
    headers = ['Period', 'Start Date', 'End Date', 'Days', 'Days in Month',
               'Beginning Book Value', 'Depreciation', 'Accumulated Depreciation', 'Ending Book Value']
    for col, header in enumerate(headers, start=1):
        cell = ws_schedule.cell(row=1, column=col, value=header)
        cell.font = header_font

    # Data rows with formulas - all values reference Summary sheet cells
    # Summary cell references:
    #   D6 = Cost, D7 = Salvage, D8 = Useful Life
    #   D11 = Depreciable Amount, D12 = Annual Depreciation, D13 = Monthly Depreciation
    #
    # Depreciation formula: Monthly_Depr × (Days / Days_in_Month)
    #   =MIN(ROUND(D13 * Days / Days_in_Month, 2), Beginning - Salvage)
    date_format = 'YYYY-MM-DD'

    for i, row in enumerate(schedule, start=2):
        ws_schedule.cell(row=i, column=1, value=row['period'])  # Period

        # Convert date strings to datetime for Excel
        start_dt = datetime.strptime(row['start_date'], '%Y-%m-%d')
        end_dt = datetime.strptime(row['end_date'], '%Y-%m-%d')

        start_cell = ws_schedule.cell(row=i, column=2, value=start_dt)
        start_cell.number_format = date_format

        end_cell = ws_schedule.cell(row=i, column=3, value=end_dt)
        end_cell.number_format = date_format

        # Days as formula: End Date - Start Date + 1
        ws_schedule.cell(row=i, column=4, value=f"=C{i}-B{i}+1")

        # Days in Month: DAY(EOMONTH(Start_Date, 0)) gives last day of month
        ws_schedule.cell(row=i, column=5, value=f"=DAY(EOMONTH(B{i},0))")

        is_first_period = (i == 2)

        # Beginning Book Value (column F)
        if is_first_period:
            ws_schedule.cell(row=i, column=6, value="=Summary!$D$6")
        else:
            ws_schedule.cell(row=i, column=6, value=f"=I{i-1}")

        # Depreciation (column G): Monthly_Depr × (Days / Days_in_Month)
        # =MIN(ROUND(Monthly_Depr * Days / Days_in_Month, 2), Beginning - Salvage)
        ws_schedule.cell(row=i, column=7,
            value=f"=MIN(ROUND(Summary!$D$13*D{i}/E{i},2),F{i}-Summary!$D$7)")

        # Accumulated Depreciation (column H)
        if is_first_period:
            ws_schedule.cell(row=i, column=8, value=f"=G{i}")
        else:
            ws_schedule.cell(row=i, column=8, value=f"=H{i-1}+G{i}")

        # Ending Book Value (column I) = Beginning - Depreciation
        ws_schedule.cell(row=i, column=9, value=f"=F{i}-G{i}")

        # Apply currency format and color-coding to value columns
        # All schedule data comes from formulas (black)
        for col in [6, 7, 8, 9]:
            cell = ws_schedule.cell(row=i, column=col)
            cell.number_format = currency_format
            if STYLES_AVAILABLE:
                cell.font = formula_style()

        # Days columns are also formulas
        if STYLES_AVAILABLE:
            ws_schedule.cell(row=i, column=4).font = formula_style()  # Days
            ws_schedule.cell(row=i, column=5).font = formula_style()  # Days in Month

    # Set column widths
    ws_schedule.column_dimensions['A'].width = 8
    ws_schedule.column_dimensions['B'].width = 12
    ws_schedule.column_dimensions['C'].width = 12
    ws_schedule.column_dimensions['D'].width = 6
    ws_schedule.column_dimensions['E'].width = 12
    ws_schedule.column_dimensions['F'].width = 18
    ws_schedule.column_dimensions['G'].width = 14
    ws_schedule.column_dimensions['H'].width = 22
    ws_schedule.column_dimensions['I'].width = 18

    wb.save(filepath)
    return True


def save_journal_xlsx(entries, filepath, asset_info, accounts):
    """Save journal entries to XLSX file with balance check formulas and color-coding.

    Color standards:
    - Blue: Hardcoded data values (amounts from schedule)
    - Black: Formula cells (balance checks, totals)
    """
    if not OPENPYXL_AVAILABLE:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Entries"

    # Styles
    header_font = Font(bold=True)
    currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    # Headers
    headers = ['Entry ID', 'Date', 'Account Code', 'Account Name',
               'Description', 'Debit', 'Credit', 'Balance Check']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font

    # Data rows
    for i, entry in enumerate(entries, start=2):
        ws.cell(row=i, column=1, value=entry['entry_id'])
        ws.cell(row=i, column=2, value=entry['date'])
        ws.cell(row=i, column=3, value=entry['account_code'])
        ws.cell(row=i, column=4, value=entry['account_name'])
        ws.cell(row=i, column=5, value=entry['description'])
        ws.cell(row=i, column=6, value=entry['debit'])
        ws.cell(row=i, column=7, value=entry['credit'])
        # Balance check formula (Debit - Credit, should be 0 for credit entries, positive for debit)
        ws.cell(row=i, column=8, value=f"=F{i}-G{i}")

        # Apply currency format and color-coding
        for col in [6, 7]:
            cell = ws.cell(row=i, column=col)
            cell.number_format = currency_format
            # Debit/Credit values are inputs (blue)
            if STYLES_AVAILABLE:
                cell.font = input_style()

        # Balance check is a formula (black)
        ws.cell(row=i, column=8).number_format = currency_format
        if STYLES_AVAILABLE:
            ws.cell(row=i, column=8).font = formula_style()

    # Add totals row - all formulas (black with bold)
    last_row = len(entries) + 2
    ws.cell(row=last_row, column=5, value="TOTALS:").font = header_font

    # Totals are formulas - apply formula_style with bold
    for col, formula in [(6, f"=SUM(F2:F{last_row-1})"),
                         (7, f"=SUM(G2:G{last_row-1})"),
                         (8, f"=F{last_row}-G{last_row}")]:
        cell = ws.cell(row=last_row, column=col, value=formula)
        cell.number_format = currency_format
        if STYLES_AVAILABLE:
            cell.font = formula_style(bold=True)
        else:
            cell.font = header_font

    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14

    wb.save(filepath)
    return True


def update_asset_register(asset_data, register_path):
    """Update or create asset register JSON file."""
    if os.path.exists(register_path):
        with open(register_path, 'r') as f:
            register = json.load(f)
    else:
        register = {'assets': [], 'metadata': {'total_assets': 0}}

    # Generate asset ID
    asset_id = f"asset_{str(uuid.uuid4())[:8]}"
    asset_data['id'] = asset_id
    asset_data['created_at'] = datetime.now().isoformat()

    register['assets'].append(asset_data)
    register['metadata']['last_updated'] = datetime.now().isoformat()
    register['metadata']['total_assets'] = len(register['assets'])

    with open(register_path, 'w') as f:
        json.dump(register, f, indent=2)

    return asset_id
