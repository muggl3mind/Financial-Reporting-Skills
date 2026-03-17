#!/usr/bin/env python3
"""
Batch Depreciation Script
Processes multiple assets and generates:
- Individual schedules (CSV + XLSX per asset)
- Consolidated schedule (all assets by period)
- PPE Rollforward (summary with Excel formulas)
- Consolidated journal entries
- Updated asset register
"""

import argparse
import json
import csv
import os
import sys
from datetime import datetime, timedelta
from collections import defaultdict

# Add shared and local modules to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LIB_PATH = os.path.join(SCRIPT_DIR, '..', '..', '..', '..', 'lib')
sys.path.insert(0, LIB_PATH)
sys.path.insert(0, SCRIPT_DIR)

from accounts import get_account_mapping, get_default_useful_life
from utils import (calculate_period_begin, get_period_dates, get_period_label,
                   VALID_PERIOD_TYPES, validate_positive_amount, validate_positive_int,
                   validate_date, get_or_create_run_timestamp)
from depreciation import (
    calculate_straight_line,
    generate_schedule,
    generate_journal_entries,
    save_schedule_csv,
    save_schedule_xlsx,
    save_journal_csv,
    save_journal_xlsx,
    update_asset_register,
    sanitize_filename,
    OPENPYXL_AVAILABLE
)

if OPENPYXL_AVAILABLE:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

# Import xlsx styles for color-coding
try:
    from xlsx_styles import input_style, formula_style, style_value_cell
    STYLES_AVAILABLE = True
except ImportError:
    STYLES_AVAILABLE = False


def parse_args():
    parser = argparse.ArgumentParser(description='Batch depreciation processing for multiple assets')
    parser.add_argument('--assets', required=True, help='JSON array of assets or path to JSON file')
    parser.add_argument('--period', required=True, choices=['monthly', 'quarterly', 'annual'],
                        help='Depreciation frequency (how often to record depreciation)')
    parser.add_argument('--output-dir', default='data/output/fixed-assets',
                        help='Output directory')
    parser.add_argument('--report-date', required=False,
                        help='Report date for journal entries and rollforward (YYYY-MM-DD). Only entries up to this date will be included.')
    parser.add_argument('--period-start', required=False,
                        help='Start of reporting period (YYYY-MM-DD). Used for PPE rollforward. Alternative to --period-type.')
    parser.add_argument('--period-type', required=False, choices=VALID_PERIOD_TYPES,
                        help='Reporting period type - auto-calculates period-start from report-date. Alternative to --period-start.')
    parser.add_argument('--run-timestamp', default=None,
                        help='Timestamp for output folder (YYYY-MM-DD_HH-MM). If not provided, uses current time.')
    return parser.parse_args()


def load_assets(assets_arg):
    """Load assets from JSON string or file path."""
    # Check if it's a file path
    if os.path.exists(assets_arg):
        with open(assets_arg, 'r') as f:
            data = json.load(f)
            # Handle both direct array and {assets: [...]} format
            return data.get('assets', data) if isinstance(data, dict) else data
    else:
        # Parse as JSON string
        return json.loads(assets_arg)


def validate_assets(assets):
    """Validate all assets in the list.

    Args:
        assets: List of asset dictionaries

    Raises:
        ValueError: If any asset has invalid values
    """
    for i, asset in enumerate(assets):
        name = asset.get('name', f'Asset #{i+1}')
        prefix = f"Asset '{name}'"

        # Validate required fields
        if 'cost' not in asset:
            raise ValueError(f"{prefix}: 'cost' is required")
        if 'purchase_date' not in asset:
            raise ValueError(f"{prefix}: 'purchase_date' is required")

        # Validate amounts
        cost = float(asset['cost'])
        validate_positive_amount(cost, f"{prefix} cost")

        salvage = float(asset.get('salvage', asset.get('salvage_value', 0)))
        validate_positive_amount(salvage, f"{prefix} salvage value", allow_zero=True)

        # Use provided useful_life or default based on category
        if 'useful_life' in asset:
            useful_life = int(asset['useful_life'])
        elif 'category' in asset:
            useful_life = get_default_useful_life(asset['category'])
        else:
            raise ValueError(f"{prefix}: Either 'useful_life' or 'category' is required")
        validate_positive_int(useful_life, f"{prefix} useful life")

        # Validate dates
        validate_date(asset['purchase_date'], f"{prefix} purchase date")


def get_all_periods(all_schedules):
    """Get a unified list of all periods across all assets."""
    all_end_dates = set()
    for asset_data in all_schedules:
        for row in asset_data['schedule']:
            all_end_dates.add(row['end_date'])
    return sorted(all_end_dates)


def save_consolidated_schedule_csv(all_schedules, filepath, report_date=None):
    """Save consolidated schedule to CSV - all assets by period.

    Shows full history from inception through report_date for balance verification.

    Args:
        all_schedules: List of asset schedule data
        filepath: Output file path
        report_date: Optional report date (YYYY-MM-DD) - only show periods through this date
    """
    all_periods = get_all_periods(all_schedules)

    # Filter periods by report_date if provided (don't show future periods)
    if report_date:
        report_date_dt = datetime.strptime(report_date, '%Y-%m-%d')
        all_periods = [p for p in all_periods if datetime.strptime(p, '%Y-%m-%d') <= report_date_dt]

    asset_names = [a['asset']['name'] for a in all_schedules]

    # Build lookup: {end_date: {asset_name: depreciation}}
    depr_lookup = defaultdict(dict)
    for asset_data in all_schedules:
        name = asset_data['asset']['name']
        for row in asset_data['schedule']:
            depr_lookup[row['end_date']][name] = row['depreciation']

    # Write CSV
    with open(filepath, 'w', newline='') as f:
        writer = csv.writer(f)

        # Header
        header = ['Period', 'End Date'] + asset_names + ['Total Depreciation', 'Total Accum Depr']
        writer.writerow(header)

        # Data rows
        running_accum = 0
        for period_num, end_date in enumerate(all_periods, start=1):
            row_data = [period_num, end_date]
            period_total = 0

            for name in asset_names:
                depr = depr_lookup[end_date].get(name, 0)
                row_data.append(f"{depr:.2f}")
                period_total += depr

            running_accum += period_total
            row_data.append(f"{period_total:.2f}")
            row_data.append(f"{running_accum:.2f}")
            writer.writerow(row_data)


def save_consolidated_schedule_xlsx(all_schedules, filepath, report_date=None):
    """Save consolidated schedule to XLSX with formulas and color-coding.

    Shows full history from inception through report_date for balance verification.

    Color standards:
    - Blue: Hardcoded depreciation values
    - Black: Formula cells (totals, running sums)

    Args:
        all_schedules: List of asset schedule data
        filepath: Output file path
        report_date: Optional report date (YYYY-MM-DD) - only show periods through this date
    """
    if not OPENPYXL_AVAILABLE:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Schedule"

    # Styles
    header_font = Font(bold=True)
    currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    all_periods = get_all_periods(all_schedules)

    # Filter periods by report_date if provided (don't show future periods)
    if report_date:
        report_date_dt = datetime.strptime(report_date, '%Y-%m-%d')
        all_periods = [p for p in all_periods if datetime.strptime(p, '%Y-%m-%d') <= report_date_dt]

    asset_names = [a['asset']['name'] for a in all_schedules]
    num_assets = len(asset_names)

    # Build lookup
    depr_lookup = defaultdict(dict)
    for asset_data in all_schedules:
        name = asset_data['asset']['name']
        for row in asset_data['schedule']:
            depr_lookup[row['end_date']][name] = row['depreciation']

    # Header
    headers = ['Period', 'End Date'] + asset_names + ['Total Depreciation', 'Total Accum Depr']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font

    # Data rows
    for row_num, (period_num, end_date) in enumerate(enumerate(all_periods, start=1), start=2):
        ws.cell(row=row_num, column=1, value=period_num)
        ws.cell(row=row_num, column=2, value=end_date)

        # Asset columns - depreciation values are inputs (blue)
        for col_offset, name in enumerate(asset_names):
            depr = depr_lookup[end_date].get(name, 0)
            cell = ws.cell(row=row_num, column=3 + col_offset, value=depr)
            cell.number_format = currency_format
            if STYLES_AVAILABLE:
                cell.font = input_style()

        # Total Depreciation formula (black)
        start_col = get_column_letter(3)
        end_col = get_column_letter(2 + num_assets)
        total_col = 3 + num_assets
        total_cell = ws.cell(row=row_num, column=total_col, value=f"=SUM({start_col}{row_num}:{end_col}{row_num})")
        total_cell.number_format = currency_format
        if STYLES_AVAILABLE:
            total_cell.font = formula_style()

        # Total Accum Depr formula (black)
        accum_col = total_col + 1
        total_col_letter = get_column_letter(total_col)
        if row_num == 2:
            accum_cell = ws.cell(row=row_num, column=accum_col, value=f"={total_col_letter}{row_num}")
        else:
            accum_col_letter = get_column_letter(accum_col)
            accum_cell = ws.cell(row=row_num, column=accum_col, value=f"={accum_col_letter}{row_num-1}+{total_col_letter}{row_num}")
        accum_cell.number_format = currency_format
        if STYLES_AVAILABLE:
            accum_cell.font = formula_style()

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    for i in range(num_assets):
        ws.column_dimensions[get_column_letter(3 + i)].width = 18
    ws.column_dimensions[get_column_letter(3 + num_assets)].width = 18
    ws.column_dimensions[get_column_letter(4 + num_assets)].width = 18

    wb.save(filepath)
    return True


def save_ppe_rollforward_xlsx(all_schedules, filepath, period_type, report_date=None, period_start=None):
    """Save PPE Rollforward schedule to XLSX with formulas and color-coding.

    Color standards:
    - Blue: Hardcoded values (beginning balances, asset costs, depreciation amounts)
    - Black: Formula cells (totals, calculated balances)
    """
    if not OPENPYXL_AVAILABLE:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "PPE Rollforward"

    # Styles
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11)
    currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    thin_border = Border(bottom=Side(style='thin'))

    # Parse dates
    report_dt = datetime.strptime(report_date, '%Y-%m-%d') if report_date else None
    period_start_dt = datetime.strptime(period_start, '%Y-%m-%d') if period_start else None

    # Calculate beginning balance date (day before period_start)
    if period_start_dt:
        beginning_date_dt = period_start_dt - timedelta(days=1)
        beginning_date_str = beginning_date_dt.strftime('%B %d, %Y')
    else:
        beginning_date_str = "Beginning"

    # Format dates for display
    if report_dt:
        report_date_str = report_dt.strftime('%B %d, %Y')
    if period_start_dt:
        period_start_str = period_start_dt.strftime('%B %d, %Y')

    # Title
    ws['A1'] = "PPE ROLLFORWARD SCHEDULE"
    ws['A1'].font = title_font
    if report_date and period_start:
        ws['A2'] = f"For the Period {period_start_str} to {report_date_str}"
        ws['A2'].font = subtitle_font
    elif report_date:
        ws['A2'] = f"Period Ending {report_date_str}"
        ws['A2'].font = subtitle_font

    # Column headers (row 4)
    ws['B4'] = "Cost"
    ws['C4'] = "Accum Depr"
    ws['D4'] = "Net Book Value"
    for col in ['B', 'C', 'D']:
        ws[f'{col}4'].font = header_font
        ws[f'{col}4'].alignment = Alignment(horizontal='right')

    # Calculate beginning balances and period activity for each asset
    asset_data_computed = []
    for asset_data in all_schedules:
        asset = asset_data['asset']
        schedule = asset_data['schedule']
        purchase_dt = datetime.strptime(asset['purchase_date'], '%Y-%m-%d')

        # Check for disposal
        disposal_date = asset.get('disposal_date')
        is_disposed = False
        disposal_accum_depr = 0
        if disposal_date and report_dt:
            disposal_dt = datetime.strptime(disposal_date, '%Y-%m-%d')
            if disposal_dt <= report_dt:
                is_disposed = True
                # Calculate accumulated depreciation through disposal
                for row in schedule:
                    row_end_dt = datetime.strptime(row['end_date'], '%Y-%m-%d')
                    if row_end_dt <= disposal_dt:
                        disposal_accum_depr += row['depreciation']
                    elif datetime.strptime(row['start_date'], '%Y-%m-%d') <= disposal_dt < row_end_dt:
                        # Prorate
                        total_days = (row_end_dt - datetime.strptime(row['start_date'], '%Y-%m-%d')).days + 1
                        days_to_disposal = (disposal_dt - datetime.strptime(row['start_date'], '%Y-%m-%d')).days + 1
                        disposal_accum_depr += round(row['depreciation'] * (days_to_disposal / total_days), 2)
                        break

        if report_dt:
            beginning_accum = 0
            period_depr = 0
            is_addition = False

            # Determine if this is an addition during the period
            if period_start_dt:
                is_addition = period_start_dt <= purchase_dt <= report_dt
            else:
                is_addition = purchase_dt <= report_dt

            for row in schedule:
                row_end_dt = datetime.strptime(row['end_date'], '%Y-%m-%d')
                row_start_dt = datetime.strptime(row['start_date'], '%Y-%m-%d')

                if period_start_dt:
                    # Beginning balance = depreciation before period_start
                    if row_end_dt < period_start_dt:
                        beginning_accum += row['depreciation']
                    # Period depreciation = depreciation within period (period_start to report_date)
                    elif row_end_dt >= period_start_dt and row_end_dt <= report_dt:
                        period_depr += row['depreciation']
                else:
                    # No period_start, use old logic
                    if row_end_dt < report_dt:
                        beginning_accum += row['depreciation']
                    elif row_end_dt == report_dt:
                        period_depr = row['depreciation']
                        break

            # Asset existed before period if it has beginning balance or was purchased before period_start
            if period_start_dt:
                asset_existed_before = purchase_dt < period_start_dt
            else:
                asset_existed_before = beginning_accum > 0 or (purchase_dt < report_dt and not is_addition)

            asset_data_computed.append({
                'asset': asset,
                'beginning_accum': beginning_accum,
                'period_depr': period_depr,
                'is_addition': is_addition,
                'asset_existed_before': asset_existed_before or is_addition,
                'is_disposed': is_disposed,
                'disposal_accum_depr': disposal_accum_depr
            })
        else:
            total_depr = sum(p['depreciation'] for p in schedule)
            asset_data_computed.append({
                'asset': asset,
                'beginning_accum': 0,
                'period_depr': total_depr,
                'is_addition': True,
                'asset_existed_before': True,
                'is_disposed': is_disposed,
                'disposal_accum_depr': disposal_accum_depr
            })

    # Row tracking
    row = 6

    # Beginning Balance
    ws[f'A{row}'] = f"BALANCE AT {beginning_date_str.upper()}"
    ws[f'A{row}'].font = header_font
    beginning_cost = sum(a['asset']['cost'] for a in asset_data_computed if a['asset_existed_before'] and not a['is_addition'])
    beginning_accum = sum(a['beginning_accum'] for a in asset_data_computed)
    ws[f'B{row}'] = beginning_cost
    ws[f'C{row}'] = -beginning_accum if beginning_accum > 0 else 0
    ws[f'D{row}'] = f"=B{row}+C{row}"
    for col in ['B', 'C']:
        ws[f'{col}{row}'].number_format = currency_format
        if STYLES_AVAILABLE:
            ws[f'{col}{row}'].font = input_style()  # Values are inputs (blue)
    ws[f'D{row}'].number_format = currency_format
    if STYLES_AVAILABLE:
        ws[f'D{row}'].font = formula_style()  # NBV is formula (black)
    beginning_row = row
    row += 2

    # Additions section
    ws[f'A{row}'] = "ADDITIONS:"
    ws[f'A{row}'].font = header_font
    row += 1

    additions_start = row
    has_additions = False
    for computed in asset_data_computed:
        if computed['is_addition']:
            has_additions = True
            asset = computed['asset']
            ws[f'A{row}'] = f"  {asset['name']}"
            ws[f'B{row}'] = asset['cost']
            ws[f'B{row}'].number_format = currency_format
            if STYLES_AVAILABLE:
                ws[f'B{row}'].font = input_style()  # Cost is input (blue)
            ws[f'C{row}'] = "—"
            ws[f'D{row}'] = f"=B{row}"
            ws[f'D{row}'].number_format = currency_format
            if STYLES_AVAILABLE:
                ws[f'D{row}'].font = formula_style()  # NBV is formula (black)
            row += 1

    if not has_additions:
        ws[f'A{row}'] = "  (none)"
        row += 1
    additions_end = row - 1

    # Total Additions - all formulas (black with bold)
    ws[f'A{row}'] = "Total Additions"
    ws[f'A{row}'].font = header_font
    if has_additions:
        ws[f'B{row}'] = f"=SUM(B{additions_start}:B{additions_end})"
    else:
        ws[f'B{row}'] = 0
    ws[f'B{row}'].number_format = currency_format
    ws[f'B{row}'].border = thin_border
    if STYLES_AVAILABLE:
        ws[f'B{row}'].font = formula_style(bold=True) if has_additions else input_style(bold=True)
    else:
        ws[f'B{row}'].font = header_font
    ws[f'C{row}'] = "—"
    if has_additions:
        ws[f'D{row}'] = f"=SUM(D{additions_start}:D{additions_end})"
    else:
        ws[f'D{row}'] = 0
    ws[f'D{row}'].number_format = currency_format
    ws[f'D{row}'].border = thin_border
    if STYLES_AVAILABLE:
        ws[f'D{row}'].font = formula_style(bold=True) if has_additions else input_style(bold=True)
    else:
        ws[f'D{row}'].font = header_font
    additions_total_row = row
    row += 2

    # Depreciation Expense section
    ws[f'A{row}'] = "DEPRECIATION EXPENSE:"
    ws[f'A{row}'].font = header_font
    row += 1

    depr_start = row
    has_depreciation = False
    for computed in asset_data_computed:
        if computed['period_depr'] > 0:
            has_depreciation = True
            asset = computed['asset']
            ws[f'A{row}'] = f"  {asset['name']}"
            ws[f'B{row}'] = "—"
            ws[f'C{row}'] = -computed['period_depr']
            ws[f'C{row}'].number_format = currency_format
            if STYLES_AVAILABLE:
                ws[f'C{row}'].font = input_style()  # Depreciation amount is input (blue)
            ws[f'D{row}'] = f"=C{row}"
            ws[f'D{row}'].number_format = currency_format
            if STYLES_AVAILABLE:
                ws[f'D{row}'].font = formula_style()  # NBV impact is formula (black)
            row += 1

    if not has_depreciation:
        ws[f'A{row}'] = "  (none)"
        row += 1
    depr_end = row - 1

    # Total Depreciation - formulas (black with bold)
    ws[f'A{row}'] = "Total Depreciation"
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'] = "—"
    if has_depreciation:
        ws[f'C{row}'] = f"=SUM(C{depr_start}:C{depr_end})"
        ws[f'D{row}'] = f"=SUM(D{depr_start}:D{depr_end})"
    else:
        ws[f'C{row}'] = 0
        ws[f'D{row}'] = 0
    ws[f'C{row}'].number_format = currency_format
    ws[f'C{row}'].border = thin_border
    ws[f'D{row}'].number_format = currency_format
    ws[f'D{row}'].border = thin_border
    if STYLES_AVAILABLE:
        ws[f'C{row}'].font = formula_style(bold=True) if has_depreciation else input_style(bold=True)
        ws[f'D{row}'].font = formula_style(bold=True) if has_depreciation else input_style(bold=True)
    else:
        ws[f'C{row}'].font = header_font
        ws[f'D{row}'].font = header_font
    depr_total_row = row
    row += 2

    # Disposals section
    ws[f'A{row}'] = "DISPOSALS:"
    ws[f'A{row}'].font = header_font
    row += 1

    disposals_start = row
    has_disposals = False
    for computed in asset_data_computed:
        if computed['is_disposed']:
            has_disposals = True
            asset = computed['asset']
            ws[f'A{row}'] = f"  {asset['name']}"
            ws[f'B{row}'] = -asset['cost']
            ws[f'B{row}'].number_format = currency_format
            if STYLES_AVAILABLE:
                ws[f'B{row}'].font = input_style()  # Cost is input (blue)
            ws[f'C{row}'] = computed['disposal_accum_depr']
            ws[f'C{row}'].number_format = currency_format
            if STYLES_AVAILABLE:
                ws[f'C{row}'].font = input_style()  # Accum depr is input (blue)
            ws[f'D{row}'] = f"=B{row}+C{row}"
            ws[f'D{row}'].number_format = currency_format
            if STYLES_AVAILABLE:
                ws[f'D{row}'].font = formula_style()  # NBV is formula (black)
            row += 1

    if not has_disposals:
        ws[f'A{row}'] = "  (none)"
        row += 1
    disposals_end = row - 1

    # Total Disposals - formulas (black with bold)
    ws[f'A{row}'] = "Total Disposals"
    ws[f'A{row}'].font = header_font
    if has_disposals:
        ws[f'B{row}'] = f"=SUM(B{disposals_start}:B{disposals_end})"
        ws[f'C{row}'] = f"=SUM(C{disposals_start}:C{disposals_end})"
        ws[f'D{row}'] = f"=SUM(D{disposals_start}:D{disposals_end})"
    else:
        ws[f'B{row}'] = 0
        ws[f'C{row}'] = 0
        ws[f'D{row}'] = 0
    for col in ['B', 'C', 'D']:
        ws[f'{col}{row}'].number_format = currency_format
        ws[f'{col}{row}'].border = thin_border
        if STYLES_AVAILABLE:
            ws[f'{col}{row}'].font = formula_style(bold=True) if has_disposals else input_style(bold=True)
        else:
            ws[f'{col}{row}'].font = header_font
    disposals_total_row = row
    row += 2

    # Ending Balance - all formulas (black with bold)
    if report_dt:
        ws[f'A{row}'] = f"BALANCE AT {report_date_str.upper()}"
    else:
        ws[f'A{row}'] = "ENDING BALANCE"
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'] = f"=B{beginning_row}+B{additions_total_row}+B{disposals_total_row}"
    ws[f'B{row}'].number_format = currency_format
    ws[f'C{row}'] = f"=C{beginning_row}+C{depr_total_row}+C{disposals_total_row}"
    ws[f'C{row}'].number_format = currency_format
    ws[f'D{row}'] = f"=B{row}+C{row}"
    ws[f'D{row}'].number_format = currency_format
    for col in ['B', 'C', 'D']:
        if STYLES_AVAILABLE:
            ws[f'{col}{row}'].font = formula_style(bold=True)
        else:
            ws[f'{col}{row}'].font = header_font
    ending_row = row
    row += 3

    # By Category section
    ws[f'A{row}'] = "BY CATEGORY:"
    ws[f'A{row}'].font = header_font
    row += 1

    categories = defaultdict(lambda: {'cost': 0, 'depr': 0})
    for computed in asset_data_computed:
        asset = computed['asset']
        cat = asset['category']
        if computed['is_disposed']:
            # Disposed assets: show category with $0 (disposal entry removed cost/accum)
            # Just ensure the category exists in the output for audit trail
            categories[cat]  # Access to create entry if not exists
        else:
            if computed['is_addition'] or computed['asset_existed_before']:
                categories[cat]['cost'] += asset['cost']
            categories[cat]['depr'] += computed['period_depr'] + computed['beginning_accum']

    for cat, data in sorted(categories.items()):
        ws[f'A{row}'] = f"  {cat}"
        ws[f'B{row}'] = data['cost']
        ws[f'B{row}'].number_format = currency_format
        if STYLES_AVAILABLE:
            ws[f'B{row}'].font = input_style()  # Cost is input (blue)
        ws[f'C{row}'] = -data['depr']
        ws[f'C{row}'].number_format = currency_format
        if STYLES_AVAILABLE:
            ws[f'C{row}'].font = input_style()  # Depr is input (blue)
        ws[f'D{row}'] = f"=B{row}+C{row}"
        ws[f'D{row}'].number_format = currency_format
        if STYLES_AVAILABLE:
            ws[f'D{row}'].font = formula_style()  # NBV is formula (black)
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18

    wb.save(filepath)
    return True


def save_consolidated_journal_csv(all_entries, filepath, report_date=None, period_start=None):
    """Save all journal entries to a single CSV.

    Args:
        all_entries: List of journal entry dicts
        filepath: Output file path
        report_date: Optional end date filter (YYYY-MM-DD)
        period_start: Optional start date filter (YYYY-MM-DD)
    """
    filtered_entries = all_entries
    if period_start:
        filtered_entries = [e for e in filtered_entries if e['date'] >= period_start]
    if report_date:
        filtered_entries = [e for e in filtered_entries if e['date'] <= report_date]

    fieldnames = ['Entry ID', 'Date', 'Asset', 'Account Code', 'Account Name',
                  'Description', 'Debit', 'Credit']

    with open(filepath, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(fieldnames)

        entry_id = 0
        for entry in filtered_entries:
            if entry['debit'] > 0:
                entry_id += 1

            writer.writerow([
                entry_id,
                entry['date'],
                entry['asset_name'],
                entry['account_code'],
                entry['account_name'],
                entry['description'],
                f"{entry['debit']:.2f}",
                f"{entry['credit']:.2f}"
            ])


def save_consolidated_journal_xlsx(all_entries, filepath, report_date=None, period_start=None):
    """Save all journal entries to XLSX with balance check formulas and color-coding.

    Color standards:
    - Blue: Hardcoded data values (debit/credit amounts)
    - Black: Formula cells (balance checks, totals)

    Args:
        all_entries: List of journal entry dicts
        filepath: Output file path
        report_date: Optional end date filter (YYYY-MM-DD)
        period_start: Optional start date filter (YYYY-MM-DD)
    """
    if not OPENPYXL_AVAILABLE:
        return False

    filtered_entries = all_entries
    if period_start:
        filtered_entries = [e for e in filtered_entries if e['date'] >= period_start]
    if report_date:
        filtered_entries = [e for e in filtered_entries if e['date'] <= report_date]

    if not filtered_entries:
        return False

    wb = Workbook()
    ws = wb.active
    if report_date:
        ws.title = f"Journal {report_date}"
    else:
        ws.title = "Journal Entries"

    header_font = Font(bold=True)
    currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    headers = ['Entry ID', 'Date', 'Asset', 'Account Code', 'Account Name',
               'Description', 'Debit', 'Credit', 'Balance Check']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font

    entry_id = 0
    for i, entry in enumerate(filtered_entries, start=2):
        if entry['debit'] > 0:
            entry_id += 1

        ws.cell(row=i, column=1, value=entry_id)
        ws.cell(row=i, column=2, value=entry['date'])
        ws.cell(row=i, column=3, value=entry['asset_name'])
        ws.cell(row=i, column=4, value=entry['account_code'])
        ws.cell(row=i, column=5, value=entry['account_name'])
        ws.cell(row=i, column=6, value=entry['description'])
        ws.cell(row=i, column=7, value=entry['debit'])
        ws.cell(row=i, column=8, value=entry['credit'])
        ws.cell(row=i, column=9, value=f"=G{i}-H{i}")

        # Apply currency format and color-coding
        for col in [7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.number_format = currency_format
            if STYLES_AVAILABLE:
                cell.font = input_style()  # Debit/Credit values are inputs (blue)

        # Balance check is a formula (black)
        ws.cell(row=i, column=9).number_format = currency_format
        if STYLES_AVAILABLE:
            ws.cell(row=i, column=9).font = formula_style()

    last_row = len(filtered_entries) + 2
    ws.cell(row=last_row, column=6, value="TOTALS:").font = header_font

    # Totals are formulas (black with bold)
    for col, formula in [(7, f"=SUM(G2:G{last_row-1})"),
                         (8, f"=SUM(H2:H{last_row-1})"),
                         (9, f"=G{last_row}-H{last_row}")]:
        cell = ws.cell(row=last_row, column=col, value=formula)
        cell.number_format = currency_format
        if STYLES_AVAILABLE:
            cell.font = formula_style(bold=True)
        else:
            cell.font = header_font

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 14

    wb.save(filepath)
    return True


def main():
    args = parse_args()

    # Prompt for report-date if not provided
    if not args.report_date:
        print("\nReport date is required to generate schedules and journal entries.")
        args.report_date = input("Enter report period end date (YYYY-MM-DD): ").strip()
        if not args.report_date:
            print("Error: Report date is required.")
            sys.exit(1)

    # Validate report date format
    try:
        from datetime import datetime
        datetime.strptime(args.report_date, '%Y-%m-%d')
    except ValueError:
        print(f"Error: Report date must be in YYYY-MM-DD format, got: {args.report_date}")
        sys.exit(1)

    # Validate report date is a quarter-end
    VALID_QUARTER_ENDS = ['03-31', '06-30', '09-30', '12-31']
    if args.report_date[5:] not in VALID_QUARTER_ENDS:
        print(f"Error: Report date must be quarter-end (Mar 31, Jun 30, Sep 30, or Dec 31).")
        print(f"       Got: {args.report_date}")
        sys.exit(1)

    # Default period-start to January 1st of report year
    report_year = args.report_date[:4]
    if args.period_start is None:
        args.period_start = f"{report_year}-01-01"
        print(f"Period: {args.period_start} to {args.report_date}")

    # Load assets
    assets = load_assets(args.assets)

    if not assets:
        print("Error: No assets provided")
        return

    # Validate all assets
    try:
        validate_assets(assets)
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)

    # Create timestamped output folder
    run_timestamp = get_or_create_run_timestamp(args.run_timestamp)
    run_output_dir = os.path.join(args.output_dir, run_timestamp)
    os.makedirs(run_output_dir, exist_ok=True)

    # Process each asset
    all_schedules = []
    all_journal_entries = []

    print(f"\n{'='*60}")
    print(f"BATCH DEPRECIATION PROCESSING")
    print(f"{'='*60}")
    print(f"\nProcessing {len(assets)} asset(s)...")

    for asset in assets:
        name = asset['name']
        category = asset['category']
        purchase_date = asset['purchase_date']
        cost = float(asset['cost'])
        salvage = float(asset.get('salvage', asset.get('salvage_value', 0)))
        # Use provided useful_life or default based on category
        if 'useful_life' in asset:
            useful_life = int(asset['useful_life'])
        else:
            useful_life = get_default_useful_life(category)
        disposal_date = asset.get('disposal_date')  # Optional disposal date

        print(f"\n  Processing: {name}")
        if disposal_date:
            print(f"    (Disposed on: {disposal_date})")

        # Get account mapping
        accounts = get_account_mapping(category)

        # Calculate depreciation
        calc_results = calculate_straight_line(cost, salvage, useful_life, args.period)

        # Generate schedule (stops at disposal_date if provided)
        schedule = generate_schedule(name, purchase_date, cost, salvage, useful_life, args.period, calc_results, disposal_date)

        # Generate journal entries (with asset name added)
        journal_entries = generate_journal_entries(name, schedule, accounts, args.period)
        for entry in journal_entries:
            entry['asset_name'] = name

        # Save individual files
        safe_name = sanitize_filename(name)

        asset_info = {
            'name': name,
            'category': category,
            'purchase_date': purchase_date,
            'cost': cost,
            'salvage': salvage,
            'useful_life': useful_life,
            'period': args.period,
            'calc': calc_results,
            'disposal_date': disposal_date  # Include if asset was disposed
        }

        # Individual schedule CSV + XLSX (go to timestamped folder)
        schedule_csv = os.path.join(run_output_dir, f'schedule_{safe_name}.csv')
        schedule_xlsx = os.path.join(run_output_dir, f'schedule_{safe_name}.xlsx')
        save_schedule_csv(schedule, schedule_csv)
        save_schedule_xlsx(schedule, schedule_xlsx, asset_info)

        # Store for consolidated outputs
        all_schedules.append({
            'asset': asset_info,
            'schedule': schedule,
            'accounts': accounts
        })
        all_journal_entries.extend(journal_entries)

        # Update asset register (stays at root level)
        register_path = os.path.join(args.output_dir, 'asset_register.json')
        asset_data = {
            'name': name,
            'category': category,
            'purchase_date': purchase_date,
            'cost': cost,
            'salvage_value': salvage,
            'useful_life_years': useful_life,
            'depreciation_method': 'straight-line',
            'reporting_period': args.period,
            'schedule_csv': f'schedule_{safe_name}.csv',
            'schedule_xlsx': f'schedule_{safe_name}.xlsx',
            'accounts': accounts,
            'calculation': calc_results
        }

        # Add disposal info if asset was disposed
        if disposal_date:
            # Calculate accumulated depreciation through disposal from schedule
            disposal_dt = datetime.strptime(disposal_date, '%Y-%m-%d')
            disposal_accum_depr = 0
            for row in schedule:
                row_end_dt = datetime.strptime(row['end_date'], '%Y-%m-%d')
                if row_end_dt <= disposal_dt:
                    disposal_accum_depr += row['depreciation']
                elif datetime.strptime(row['start_date'], '%Y-%m-%d') <= disposal_dt < row_end_dt:
                    # Prorate for partial period
                    total_days = (row_end_dt - datetime.strptime(row['start_date'], '%Y-%m-%d')).days + 1
                    days_to_disposal = (disposal_dt - datetime.strptime(row['start_date'], '%Y-%m-%d')).days + 1
                    disposal_accum_depr += round(row['depreciation'] * (days_to_disposal / total_days), 2)

            book_value = cost - disposal_accum_depr
            proceeds = float(asset.get('disposal_proceeds', 0))
            gain_loss = proceeds - book_value

            asset_data['disposal_date'] = disposal_date
            asset_data['disposal_info'] = {
                'proceeds': proceeds,
                'accumulated_depreciation': disposal_accum_depr,
                'book_value': book_value,
                'gain_loss': gain_loss,
                'is_gain': gain_loss >= 0
            }
            asset_data['status'] = 'disposed'

        update_asset_register(asset_data, register_path)

        print(f"    - Schedule: {len(schedule)} periods")
        print(f"    - Total depreciation: ${sum(p['depreciation'] for p in schedule):,.2f}")

    # Generate consolidated outputs
    report_date = args.report_date

    # Calculate period_start using shared utilities
    if args.period_start:
        period_start = args.period_start
        period_label = f"{period_start} to {report_date}"
    elif args.period_type and report_date:
        period_start = calculate_period_begin(report_date, args.period_type)
        period_label = get_period_label(period_start, report_date, args.period_type)
    elif report_date:
        period_start = f"{report_date[:4]}-01-01"  # Default to Jan 1 of report year
        period_label = f"FY {report_date[:4]} ({period_start} to {report_date})"
    else:
        period_start = None
        period_label = None

    print(f"\nGenerating consolidated outputs...")
    if period_label:
        print(f"  Report period: {period_label}")

    # Consolidated Schedule (shows inception to date for balance verification)
    consolidated_schedule_csv = os.path.join(run_output_dir, 'consolidated_schedule.csv')
    consolidated_schedule_xlsx = os.path.join(run_output_dir, 'consolidated_schedule.xlsx')
    save_consolidated_schedule_csv(all_schedules, consolidated_schedule_csv, report_date)
    save_consolidated_schedule_xlsx(all_schedules, consolidated_schedule_xlsx, report_date)
    print(f"  - Consolidated schedule: {consolidated_schedule_csv}")

    # PPE Rollforward
    if report_date:
        ppe_rollforward_xlsx = os.path.join(run_output_dir, f'ppe_rollforward_{report_date}.xlsx')
    else:
        ppe_rollforward_xlsx = os.path.join(run_output_dir, 'ppe_rollforward.xlsx')
    save_ppe_rollforward_xlsx(all_schedules, ppe_rollforward_xlsx, args.period, report_date, period_start)
    print(f"  - PPE Rollforward: {ppe_rollforward_xlsx}")

    # Consolidated Journal
    if report_date:
        consolidated_journal_csv = os.path.join(run_output_dir, f'journal_{report_date}.csv')
        consolidated_journal_xlsx = os.path.join(run_output_dir, f'journal_{report_date}.xlsx')
    else:
        consolidated_journal_csv = os.path.join(run_output_dir, 'consolidated_journal.csv')
        consolidated_journal_xlsx = os.path.join(run_output_dir, 'consolidated_journal.xlsx')
    save_consolidated_journal_csv(all_journal_entries, consolidated_journal_csv, report_date, period_start)
    save_consolidated_journal_xlsx(all_journal_entries, consolidated_journal_xlsx, report_date, period_start)
    print(f"  - Consolidated journal: {consolidated_journal_csv}")

    # Summary
    total_cost = sum(a['asset']['cost'] for a in all_schedules)
    if report_date:
        total_depr = 0
        for a in all_schedules:
            for p in a['schedule']:
                if p['end_date'] == report_date:
                    total_depr += p['depreciation']
    else:
        total_depr = sum(sum(p['depreciation'] for p in a['schedule']) for a in all_schedules)

    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"Assets processed: {len(assets)}")
    print(f"Total cost: ${total_cost:,.2f}")
    if report_date:
        print(f"Period depreciation ({report_date}): ${total_depr:,.2f}")
    else:
        print(f"Total depreciation (full lifecycle): ${total_depr:,.2f}")
    print(f"\nOutput files:")
    print(f"  Individual schedules: schedule_<asset>.csv/.xlsx")
    print(f"  Consolidated schedule: consolidated_schedule.csv/.xlsx")
    if report_date:
        print(f"  PPE Rollforward: ppe_rollforward_{report_date}.xlsx")
        print(f"  Journal entries: journal_{report_date}.csv/.xlsx")
    else:
        print(f"  PPE Rollforward: ppe_rollforward.xlsx")
        print(f"  Journal entries: consolidated_journal.csv/.xlsx")
    print(f"  Asset register: asset_register.json")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    main()
