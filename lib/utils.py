#!/usr/bin/env python3
"""
Shared utility functions for depreciation and securities accounting.
"""

import csv
import os
from datetime import datetime
from typing import Tuple, Optional


# =============================================================================
# INPUT VALIDATION UTILITIES
# =============================================================================

def validate_positive_amount(value, field_name, allow_zero=False):
    """Validate that a numeric amount is positive (or optionally zero).

    Args:
        value: The numeric value to validate
        field_name: Name of field for error messages
        allow_zero: If True, zero is allowed; if False, must be strictly positive

    Raises:
        ValueError: If value is negative or (if allow_zero=False) zero
    """
    if value < 0:
        raise ValueError(f"{field_name} cannot be negative (got: {value})")
    if not allow_zero and value == 0:
        raise ValueError(f"{field_name} must be greater than zero (got: {value})")


def validate_positive_int(value, field_name, allow_zero=False):
    """Validate that an integer is positive (or optionally zero).

    Args:
        value: The integer value to validate
        field_name: Name of field for error messages
        allow_zero: If True, zero is allowed; if False, must be strictly positive

    Raises:
        ValueError: If value is negative or (if allow_zero=False) zero
    """
    if value < 0:
        raise ValueError(f"{field_name} cannot be negative (got: {value})")
    if not allow_zero and value == 0:
        raise ValueError(f"{field_name} must be greater than zero (got: {value})")


# =============================================================================
# OUTPUT FILE UTILITIES
# =============================================================================

def get_run_timestamp():
    """Generate timestamp string for current run folder.

    Returns:
        String in format 'YYYY-MM-DD_HH-MM' for use as folder name.
        Scripts run within the same minute share the same folder.
    """
    return datetime.now().strftime('%Y-%m-%d_%H-%M')


def get_or_create_run_timestamp(provided_timestamp=None):
    """Get provided timestamp or generate a new one.

    Use this when processing batches - pass the same timestamp to all scripts
    so they output to the same folder.

    Args:
        provided_timestamp: Optional timestamp string (YYYY-MM-DD_HH-MM format)

    Returns:
        The provided timestamp if valid, or a newly generated one.
    """
    if provided_timestamp:
        # Basic validation - should match YYYY-MM-DD_HH-MM format
        import re
        if re.match(r'^\d{4}-\d{2}-\d{2}_\d{2}-\d{2}$', provided_timestamp):
            return provided_timestamp
    return get_run_timestamp()


# =============================================================================
# DATE AND PERIOD UTILITIES
# =============================================================================

def validate_date(date_str, field_name="Date"):
    """Validate date string is in YYYY-MM-DD format and return datetime object.

    Args:
        date_str: Date string to validate
        field_name: Name of field for error messages

    Returns:
        datetime object if valid

    Raises:
        ValueError: If date format is invalid
    """
    if not date_str:
        raise ValueError(f"{field_name} is required")
    try:
        return datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        raise ValueError(f"{field_name} must be in YYYY-MM-DD format, got: {date_str}")


def calculate_period_begin(report_date: str, period_type: str) -> str:
    """Calculate period begin date from report date and period type.

    Args:
        report_date: Period end date (YYYY-MM-DD)
        period_type: 'quarterly', 'half-year', or 'annual'

    Returns:
        Period begin date (YYYY-MM-DD)

    Example:
        >>> calculate_period_begin('2025-09-30', 'quarterly')
        '2025-07-01'
        >>> calculate_period_begin('2025-09-30', 'annual')
        '2025-01-01'
    """
    report_dt = datetime.strptime(report_date, '%Y-%m-%d')
    year = report_dt.year
    month = report_dt.month

    if period_type == 'quarterly':
        # Q1: Jan 1 - Mar 31, Q2: Apr 1 - Jun 30, Q3: Jul 1 - Sep 30, Q4: Oct 1 - Dec 31
        if month <= 3:
            begin_dt = datetime(year, 1, 1)
        elif month <= 6:
            begin_dt = datetime(year, 4, 1)
        elif month <= 9:
            begin_dt = datetime(year, 7, 1)
        else:
            begin_dt = datetime(year, 10, 1)
    elif period_type == 'half-year':
        # H1: Jan 1 - Jun 30, H2: Jul 1 - Dec 31
        if month <= 6:
            begin_dt = datetime(year, 1, 1)
        else:
            begin_dt = datetime(year, 7, 1)
    elif period_type == 'annual':
        begin_dt = datetime(year, 1, 1)
    else:
        raise ValueError(f"Invalid period_type: {period_type}. Must be 'quarterly', 'half-year', or 'annual'")

    return begin_dt.strftime('%Y-%m-%d')


def get_period_dates(report_date: str, period_begin: Optional[str] = None,
                     period_type: Optional[str] = None) -> Tuple[str, str]:
    """Get period begin and end dates, validating inputs.

    Must provide either period_begin OR period_type.

    Args:
        report_date: Period end date (YYYY-MM-DD)
        period_begin: Explicit period begin date (YYYY-MM-DD)
        period_type: 'quarterly', 'half-year', or 'annual'

    Returns:
        Tuple of (period_begin, report_date) as strings

    Raises:
        ValueError: If neither period_begin nor period_type provided
    """
    if not period_begin and not period_type:
        raise ValueError("Either period_begin or period_type is required")

    if period_begin:
        # Validate the date format
        validate_date(period_begin, "period_begin")
        return (period_begin, report_date)
    else:
        calculated_begin = calculate_period_begin(report_date, period_type)
        return (calculated_begin, report_date)


def get_period_label(period_begin: str, report_date: str, period_type: Optional[str] = None) -> str:
    """Get a human-readable label for the period.

    Args:
        period_begin: Period begin date (YYYY-MM-DD)
        report_date: Period end date (YYYY-MM-DD)
        period_type: Optional period type for enhanced label

    Returns:
        Period label string

    Example:
        >>> get_period_label('2025-07-01', '2025-09-30', 'quarterly')
        'Q3 2025 (2025-07-01 to 2025-09-30)'
    """
    report_dt = datetime.strptime(report_date, '%Y-%m-%d')
    year = report_dt.year
    month = report_dt.month

    if period_type == 'quarterly':
        if month <= 3:
            quarter = 'Q1'
        elif month <= 6:
            quarter = 'Q2'
        elif month <= 9:
            quarter = 'Q3'
        else:
            quarter = 'Q4'
        return f"{quarter} {year} ({period_begin} to {report_date})"
    elif period_type == 'half-year':
        half = 'H1' if month <= 6 else 'H2'
        return f"{half} {year} ({period_begin} to {report_date})"
    elif period_type == 'annual':
        return f"FY {year} ({period_begin} to {report_date})"
    else:
        return f"{period_begin} to {report_date}"


VALID_PERIOD_TYPES = ['quarterly', 'half-year', 'annual']


# =============================================================================
# FILE AND FORMATTING UTILITIES
# =============================================================================

# Try to import openpyxl for Excel output
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import xlsx styles for color-coding
try:
    from xlsx_styles import input_style, formula_style
    STYLES_AVAILABLE = True
except ImportError:
    STYLES_AVAILABLE = False

# Common formatting
CURRENCY_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


def sanitize_filename(name):
    """Convert name to safe filename."""
    return name.lower().replace(' ', '_').replace('-', '_').replace('.', '').replace(',', '')


def save_journal_csv(entries, filepath):
    """Save journal entries to CSV file.

    Args:
        entries: List of dicts with keys: date, account_code, account_name,
                 description, debit, credit, and optionally asset_name/type
        filepath: Output file path
    """
    # Determine if we have asset_name or type columns
    has_asset = any('asset_name' in e or 'asset' in e for e in entries)
    has_type = any('type' in e for e in entries)

    if has_asset and has_type:
        fieldnames = ['Date', 'Type', 'Asset', 'Account Code', 'Account Name',
                      'Description', 'Debit', 'Credit']
    elif has_asset:
        fieldnames = ['Date', 'Asset', 'Account Code', 'Account Name',
                      'Description', 'Debit', 'Credit']
    else:
        fieldnames = ['Date', 'Account Code', 'Account Name',
                      'Description', 'Debit', 'Credit']

    with open(filepath, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(fieldnames)

        for entry in entries:
            row = [entry['date']]

            if has_type:
                row.append(entry.get('type', ''))
            if has_asset:
                row.append(entry.get('asset_name', entry.get('asset', '')))

            row.extend([
                entry['account_code'],
                entry['account_name'],
                entry['description'],
                f"{entry['debit']:.2f}" if entry['debit'] > 0 else '',
                f"{entry['credit']:.2f}" if entry['credit'] > 0 else ''
            ])
            writer.writerow(row)


def save_journal_xlsx(entries, filepath, title="Journal Entries"):
    """Save journal entries to XLSX file with totals and color-coding.

    Color standards:
    - Blue: Hardcoded data values (debit/credit amounts)
    - Black: Formula cells (totals)

    Args:
        entries: List of journal entry dicts
        filepath: Output file path
        title: Sheet title

    Returns:
        True if successful, False if openpyxl not available
    """
    if not OPENPYXL_AVAILABLE:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    header_font = Font(bold=True)

    # Determine columns
    has_asset = any('asset_name' in e or 'asset' in e for e in entries)
    has_type = any('type' in e for e in entries)

    if has_asset and has_type:
        headers = ['Date', 'Type', 'Asset', 'Account Code', 'Account Name',
                   'Description', 'Debit', 'Credit']
        debit_col = 7
        credit_col = 8
    elif has_asset:
        headers = ['Date', 'Asset', 'Account Code', 'Account Name',
                   'Description', 'Debit', 'Credit']
        debit_col = 6
        credit_col = 7
    else:
        headers = ['Date', 'Account Code', 'Account Name',
                   'Description', 'Debit', 'Credit']
        debit_col = 5
        credit_col = 6

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font

    # Write data
    for i, entry in enumerate(entries, start=2):
        col = 1
        ws.cell(row=i, column=col, value=entry['date'])
        col += 1

        if has_type:
            ws.cell(row=i, column=col, value=entry.get('type', ''))
            col += 1
        if has_asset:
            ws.cell(row=i, column=col, value=entry.get('asset_name', entry.get('asset', '')))
            col += 1

        ws.cell(row=i, column=col, value=entry['account_code'])
        col += 1
        ws.cell(row=i, column=col, value=entry['account_name'])
        col += 1
        ws.cell(row=i, column=col, value=entry['description'])
        col += 1

        # Debit/Credit are input values (blue)
        if entry['debit'] > 0:
            cell = ws.cell(row=i, column=col, value=entry['debit'])
            cell.number_format = CURRENCY_FORMAT
            if STYLES_AVAILABLE:
                cell.font = input_style()
        col += 1

        if entry['credit'] > 0:
            cell = ws.cell(row=i, column=col, value=entry['credit'])
            cell.number_format = CURRENCY_FORMAT
            if STYLES_AVAILABLE:
                cell.font = input_style()

    # Add totals row - formulas (black with bold)
    total_row = len(entries) + 2
    ws.cell(row=total_row, column=debit_col - 1, value="TOTALS:").font = header_font

    debit_col_letter = chr(64 + debit_col)
    credit_col_letter = chr(64 + credit_col)

    # Totals are formulas (black with bold)
    debit_total = ws.cell(row=total_row, column=debit_col, value=f"=SUM({debit_col_letter}2:{debit_col_letter}{total_row-1})")
    debit_total.number_format = CURRENCY_FORMAT
    if STYLES_AVAILABLE:
        debit_total.font = formula_style(bold=True)
    else:
        debit_total.font = header_font

    credit_total = ws.cell(row=total_row, column=credit_col, value=f"=SUM({credit_col_letter}2:{credit_col_letter}{total_row-1})")
    credit_total.number_format = CURRENCY_FORMAT
    if STYLES_AVAILABLE:
        credit_total.font = formula_style(bold=True)
    else:
        credit_total.font = header_font

    # Set column widths
    ws.column_dimensions['A'].width = 12
    if has_type:
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
    elif has_asset:
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
    else:
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14

    wb.save(filepath)
    return True
