#!/usr/bin/env python3
"""
XLSX Styling Utilities

Centralized color-coding standards for financial spreadsheets.
Based on industry-standard financial modeling conventions.

Color Standards:
- Blue: Hardcoded inputs (user-changeable values)
- Black: Formulas and calculations
- Green: Internal worksheet links
- Red: External file links
- Yellow fill: Key assumptions
"""

try:
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# =============================================================================
# COLOR CONSTANTS (ARGB format for openpyxl)
# =============================================================================

COLORS = {
    'INPUT': 'FF0000FF',         # Blue - hardcoded values
    'FORMULA': 'FF000000',       # Black - calculated cells
    'LINK_INTERNAL': 'FF008000', # Green - sheet references
    'LINK_EXTERNAL': 'FFFF0000', # Red - external file links
    'ASSUMPTION_FILL': 'FFFFFF00',  # Yellow - key assumptions
}


# =============================================================================
# STYLE FACTORY FUNCTIONS
# =============================================================================

def input_style(bold=False):
    """Return Font style for hardcoded input values (blue).

    Use for: cost, salvage value, purchase price, proceeds, dates entered by user.

    Args:
        bold: If True, make the font bold

    Returns:
        openpyxl Font object with blue color
    """
    if not OPENPYXL_AVAILABLE:
        return None
    return Font(color=COLORS['INPUT'], bold=bold)


def formula_style(bold=False):
    """Return Font style for formula/calculated cells (black).

    Use for: depreciation calculations, totals, balances derived from formulas.

    Args:
        bold: If True, make the font bold

    Returns:
        openpyxl Font object with black color
    """
    if not OPENPYXL_AVAILABLE:
        return None
    return Font(color=COLORS['FORMULA'], bold=bold)


def link_internal_style(bold=False):
    """Return Font style for internal worksheet links (green).

    Use for: References to other sheets within the same workbook.

    Args:
        bold: If True, make the font bold

    Returns:
        openpyxl Font object with green color
    """
    if not OPENPYXL_AVAILABLE:
        return None
    return Font(color=COLORS['LINK_INTERNAL'], bold=bold)


def link_external_style(bold=False):
    """Return Font style for external file links (red).

    Use for: References to external workbooks or files.

    Args:
        bold: If True, make the font bold

    Returns:
        openpyxl Font object with red color
    """
    if not OPENPYXL_AVAILABLE:
        return None
    return Font(color=COLORS['LINK_EXTERNAL'], bold=bold)


def assumption_fill():
    """Return PatternFill for key assumptions (yellow background).

    Use for: Critical assumptions that auditors/reviewers should verify.

    Returns:
        openpyxl PatternFill object with yellow background
    """
    if not OPENPYXL_AVAILABLE:
        return None
    return PatternFill(start_color=COLORS['ASSUMPTION_FILL'],
                       end_color=COLORS['ASSUMPTION_FILL'],
                       fill_type='solid')


def header_style():
    """Return Font style for headers (bold black).

    Returns:
        openpyxl Font object with bold black
    """
    if not OPENPYXL_AVAILABLE:
        return None
    return Font(bold=True, color=COLORS['FORMULA'])


# =============================================================================
# CELL STYLING HELPERS
# =============================================================================

def apply_cell_style(cell, style_type, bold=False):
    """Apply appropriate style to a cell based on its content type.

    Args:
        cell: openpyxl cell object
        style_type: One of 'input', 'formula', 'link_internal', 'link_external',
                    'assumption', 'header'
        bold: If True, make the font bold (ignored for 'assumption' and 'header')
    """
    if not OPENPYXL_AVAILABLE:
        return

    style_map = {
        'input': lambda: input_style(bold),
        'formula': lambda: formula_style(bold),
        'link_internal': lambda: link_internal_style(bold),
        'link_external': lambda: link_external_style(bold),
        'header': lambda: header_style(),
    }

    if style_type == 'assumption':
        cell.fill = assumption_fill()
        cell.font = input_style(bold)
    elif style_type in style_map:
        cell.font = style_map[style_type]()


def style_value_cell(cell, is_formula=False, bold=False):
    """Style a cell based on whether it contains a formula or hardcoded value.

    This is a convenience function for the common case of styling numeric cells.

    Args:
        cell: openpyxl cell object
        is_formula: True if the cell contains a formula (will be styled black)
        bold: If True, make the font bold
    """
    if is_formula:
        apply_cell_style(cell, 'formula', bold)
    else:
        apply_cell_style(cell, 'input', bold)


def is_formula_value(value):
    """Check if a cell value is a formula.

    Args:
        value: Cell value to check

    Returns:
        True if value is a string starting with '='
    """
    return isinstance(value, str) and value.startswith('=')
